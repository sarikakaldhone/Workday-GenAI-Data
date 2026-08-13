# -*- coding: utf-8 -*-
#!/usr/bin/env python3
"""
DGW Validation Engine -- Single-file executable
Run: python DGW_Validator.py
Then open: http://localhost:5000
"""

# -- Standard imports --------------------------------------------------------
import os
import sys
import re
import uuid
import threading
import datetime
import shutil
import webbrowser
from typing import Callable, Dict, List, Optional, Set, Tuple

# -- Third-party imports -----------------------------------------------------
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    from flask import Flask, render_template_string, request, send_file, jsonify, Response
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Install with: pip install flask openpyxl")
    sys.exit(1)


# =======================================================================
# SECTION 1: RULES PARSER  (from rules_parser.py)
# =======================================================================

def parse_rules_workbook(rules_path: str) -> dict:
    """
    Schema-driven rules engine. Reads ALL validation rules from the Excel
    rules workbook at runtime — no code changes needed when rules change.

    For each 'X Validation Rules' sheet:
      - If a 'Rules Table' block is found, structured rules are parsed from it.
      - Otherwise, legacy format is parsed (Required Fields + free-text interpretation).

    For each 'X Reference ID' sheet, allowed_values are built as before.

    Returns: {
      rules_base_key: {
        'rules': [rule_dicts],        # list of rule dicts
        'allowed_values': {col: [values]},
        'matched_section': str,
      }
    }
    """
    wb = load_workbook(rules_path, data_only=True)

    rules_by_base: dict = {}
    refs_by_base:  dict = {}

    for ws in wb.worksheets:
        sname   = ws.title
        sname_l = sname.lower()

        if _is_validation_rules_sheet(sname_l):
            base = _extract_base(sname_l)
            # Try structured Rules Table first
            rules_list = _parse_rules_table(ws, sname)
            if rules_list is not None:
                # Structured Rules Table found
                rules_by_base[base] = {
                    'rules': rules_list,
                    'matched_section': sname,
                    '_has_rules_table': True,
                }
            else:
                # Legacy: Required Fields section + free-text interpretation
                required = _parse_required_fields(ws)
                free_text_rules = _extract_free_text_rules(ws, sname)
                legacy_rules = [_make_rule(rf, 'Required', source=f'Legacy: {sname}')
                                for rf in required]
                legacy_rules.extend(free_text_rules)
                rules_by_base[base] = {
                    'rules': legacy_rules,
                    'matched_section': sname,
                    '_has_rules_table': False,
                }

        elif _is_reference_sheet(sname_l):
            base = _extract_base(sname_l)
            allowed = _parse_reference_id(ws)
            if allowed:
                refs_by_base[base] = allowed

    wb.close()

    # Merge rules + refs by base key into combined result
    all_bases = set(rules_by_base) | set(refs_by_base)
    result = {}
    for base in all_bases:
        rules_entry  = rules_by_base.get(base, {})
        ref_allowed  = refs_by_base.get(base, {})
        existing_rules = list(rules_entry.get('rules', []))
        has_rules_table = rules_entry.get('_has_rules_table', False)

        # For legacy sheets: auto-add AllowedValues rules from Reference ID columns
        # that are not already covered by an explicit rule in the sheet.
        if not has_rules_table:
            covered_av = {r['field'] for r in existing_rules
                          if r['rule_type'].lower() == 'allowedvalues'}
            for col in ref_allowed:
                if col not in covered_av:
                    existing_rules.append(_make_rule(
                        col, 'AllowedValues',
                        source='Legacy: Reference ID sheet',
                    ))

        result[base] = {
            'rules':          existing_rules,
            'allowed_values': ref_allowed,
            'matched_section': rules_entry.get('matched_section', base),
        }

    return result


# Semantic fragment hints: when a DGW sheet name contains a given
# keyword, search the rules cache for a key that contains the hint.
# This handles abbreviated base keys (e.g. 'ccw', 'emp comp change')
# that can never be matched by word overlap alone.
# Format: (dgw_name_must_contain, cache_key_must_contain)
# Add new rows here when a rules workbook uses abbreviated sheet names.
_ABBREV_HINTS = [
    ('compensation',   'comp'),       # 'Employee Compensation Data' -> 'emp comp change'
    ('contingent',     'ccw'),        # 'Contract Contingent Worker' -> 'ccw'
    ('pay group',      'paygrp'),     # extra alias
    ('paygroup',       'paygrp'),     # extra alias
]


def _match_rules_for_sheet(dgw_sheet_name: str, rules_cache: dict) -> dict:
    """
    Find the best matching rules entry for a DGW sheet name.

    Priority:
      1. Exact base match
      2. Semantic abbreviation hint  (dgw contains keyword -> cache key contains hint)
      3. Substring match either direction
      4. Word-level overlap (>= 2 shared words to avoid false positives)
      5. Empty fallback -- engine uses DGW row-5 metadata instead
    """
    sn      = dgw_sheet_name.strip().lower()
    sn_base = _extract_base(sn)
    _empty  = {'rules': [], 'allowed_values': {}, 'matched_section': ''}

    # 1. Exact base match
    for candidate in (sn_base, sn):
        if candidate in rules_cache:
            return rules_cache[candidate]

    # 2. Abbreviation hint: DGW name contains keyword AND cache key contains hint
    for dgw_kw, cache_hint in _ABBREV_HINTS:
        if dgw_kw in sn:
            for ck, entry in rules_cache.items():
                if cache_hint in ck:
                    return entry

    # 3. Substring: full DGW name contains a cache base key or vice-versa
    for base, entry in rules_cache.items():
        if base in sn or sn in base:
            return entry

    # 4. Word-level overlap -- require >= 2 to avoid weak matches like 'employee'
    sn_words = set(sn.split())
    best_score, best_entry = 0, None
    for base, entry in rules_cache.items():
        overlap = len(sn_words & set(base.split()))
        if overlap > best_score:
            best_score = overlap
            best_entry = entry
    if best_score >= 2 and best_entry is not None:
        return best_entry

    return _empty


def _is_validation_rules_sheet(name_l: str) -> bool:
    return ('validation rule' in name_l) and ('reference' not in name_l)


def _is_reference_sheet(name_l: str) -> bool:
    return 'reference' in name_l and 'id' in name_l


def _extract_base(name_l: str) -> str:
    for suffix in ('validation rules', 'validation rule', 'reference id', 'reference'):
        name_l = name_l.replace(suffix, '').strip()
    return name_l.strip()


def _best_match(key: str, d: dict) -> Optional[str]:
    if not d:
        return None
    key_l = key.lower()
    if key_l in d:
        return key_l
    for k in d:
        if key_l in k or k in key_l:
            return k
    key_words = set(key_l.split())
    best_score, best_k = 0, None
    for k in d:
        overlap = len(key_words & set(k.split()))
        if overlap > best_score:
            best_score, best_k = overlap, k
    return best_k if best_score > 0 else None


def _parse_required_fields(ws) -> list:
    required = []
    in_required_section = False
    for r in range(1, (ws.max_row or 1) + 1):
        col_a = str(ws.cell(row=r, column=1).value or '').strip()
        col_b = str(ws.cell(row=r, column=2).value or '').strip()
        col_a_l = col_a.lower()

        # Accept typos like "Required fileds", "Required Field" — 2-word section header
        if col_a_l.startswith('required') and len(col_a.split()) <= 3:
            in_required_section = True
            if col_b:
                required.append(col_b)
            continue

        if col_a_l.startswith('validation'):
            break

        if in_required_section and col_b:
            required.append(col_b)

    return required


def _make_rule(field: str, rule_type: str, rule_value: str = '', severity: str = 'High',
               error_message: str = '', condition_field: str = '', condition_value: str = '',
               source: str = '') -> dict:
    """Create a normalised rule dict."""
    return {
        'field': field, 'rule_type': rule_type, 'rule_value': rule_value,
        'severity': severity or 'High', 'error_message': error_message,
        'condition_field': condition_field, 'condition_value': condition_value,
        'source': source,
    }


def _parse_rules_table(ws, sname: str) -> Optional[list]:
    """
    Scan the sheet top-down for a 'Rules Table' marker row (col A value).
    If found, parse all subsequent rows as structured rules.
    Returns list of rule dicts, or None if no Rules Table marker exists.
    """
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    rules_table_row = None

    for r in range(1, max_row + 1):
        col_a = str(ws.cell(row=r, column=1).value or '').strip().lower()
        if col_a == 'rules table':
            rules_table_row = r
            break

    if rules_table_row is None:
        return None  # No Rules Table — caller uses legacy path

    header_row = rules_table_row + 1
    if header_row > max_row:
        return []

    # Map logical column names using case-insensitive substring matching
    col_map: dict = {}
    for c in range(1, max_col + 1):
        h = str(ws.cell(row=header_row, column=c).value or '').strip().lower()
        if not h:
            continue
        if 'field' in h and 'condition' not in h:
            col_map.setdefault('field', c)
        elif 'rule type' in h or h == 'type':
            col_map.setdefault('rule_type', c)
        elif ('rule value' in h or h == 'value') and 'condition' not in h:
            col_map.setdefault('rule_value', c)
        elif 'severity' in h:
            col_map.setdefault('severity', c)
        elif ('error' in h or 'message' in h) and 'condition' not in h:
            col_map.setdefault('error_message', c)
        elif 'condition field' in h:
            col_map.setdefault('condition_field', c)
        elif 'condition value' in h:
            col_map.setdefault('condition_value', c)

    rules = []
    for r in range(header_row + 1, max_row + 1):
        field_val = str(ws.cell(row=r, column=col_map.get('field', 1)).value or '').strip()
        if not field_val:
            break  # Blank field column signals end of rules table

        rule_type  = str(ws.cell(row=r, column=col_map.get('rule_type', 2)).value or '').strip()
        rule_value = str(ws.cell(row=r, column=col_map.get('rule_value', 3)).value or '').strip()
        severity   = str(ws.cell(row=r, column=col_map.get('severity', 4)).value or '').strip() or 'High'
        error_msg  = str(ws.cell(row=r, column=col_map.get('error_message', 5)).value or '').strip()
        cond_field = str(ws.cell(row=r, column=col_map.get('condition_field', 6)).value or '').strip()
        cond_val   = str(ws.cell(row=r, column=col_map.get('condition_value', 7)).value or '').strip()

        if not rule_type:
            continue

        rules.append(_make_rule(
            field=field_val, rule_type=rule_type, rule_value=rule_value,
            severity=severity, error_message=error_msg,
            condition_field=cond_field, condition_value=cond_val,
            source=f'{sname} Row {r}',
        ))

    return rules


def _titlecase_field(s: str) -> str:
    return ' '.join(w.capitalize() for w in s.split())


def _extract_free_text_rules(ws, sname: str) -> list:
    """
    Scan for the 'Validation Rules' free-text block and interpret rules.
    Reads whichever column (A or B) contains the rule text — some sheets
    put text in col B, others put it entirely in col A.
    """
    max_row = ws.max_row or 1
    in_vr_section = False
    text_lines = []
    for r in range(1, max_row + 1):
        col_a = str(ws.cell(row=r, column=1).value or '').strip()
        col_b = str(ws.cell(row=r, column=2).value or '').strip()
        if col_a.lower().startswith('validation'):
            in_vr_section = True
            if col_b:
                text_lines.append(col_b)
            continue
        if in_vr_section:
            # Prefer col_b; fall back to col_a when col_b is blank
            text = col_b or col_a
            if text:
                text_lines.append(text)
    return _interpret_free_text_rules(text_lines, sname)


def _interpret_free_text_rules(text_rules: list, section_name: str) -> list:
    """Extract structured rules from free-text validation rule descriptions."""
    rules = []
    for text in text_rules:
        t = text.lower().strip()

        # Unique: "X should be unique" / "X must be unique"
        m = re.search(r'(\w[\w\s]+?)\s+(?:should be|must be)\s+unique', t)
        if m:
            field = _titlecase_field(m.group(1).strip())
            rules.append(_make_rule(field, 'Unique', source=f'Legacy: {section_name}'))

        # Numeric: "X can only be numeric"
        m = re.search(r'(\w[\w\s]+?)\s+can only be numeric', t)
        if m:
            rules.append(_make_rule(_titlecase_field(m.group(1).strip()), 'Numeric',
                                    source=f'Legacy: {section_name}'))

        # DateRange: "X can't/cannot be earlier/greater than Y"
        m = re.search(
            r"(.+?)\s+can(?:'t| not)\s+be\s+(?:earlier|greater)\s+(?:to|than)\s+(.+?)(?:\s+date)?\s*$", t)
        if m:
            end_field   = _titlecase_field(m.group(1).strip())
            start_field = _titlecase_field(m.group(2).strip())
            rules.append(_make_rule(end_field, 'DateRange', rule_value=start_field,
                                    source=f'Legacy: {section_name}'))

        # Regex: "X should begin with 'Y'" / "X must start with 'Y'"
        # Match against original text to preserve the prefix case (e.g. "GRA" not "gra")
        m = re.search(
            r'(\w[\w\s]+?)\s+(?:should|must)\s+(?:begin|start)\s+with\s+["\']?([^"\'",]+)["\']?',
            text, re.IGNORECASE)
        if m:
            field  = _titlecase_field(m.group(1).strip())
            prefix = m.group(2).strip()
            # Honour "case insensitive" / "case unsensitive" / "case-insensitive" qualifier
            case_insensitive = bool(re.search(r'case[- ]?(?:in|un)sensitive', text, re.IGNORECASE))
            pattern = f'(?i)^{re.escape(prefix)}' if case_insensitive else f'^{re.escape(prefix)}'
            rules.append(_make_rule(field, 'Regex', rule_value=pattern,
                                    error_message=f'{field} must begin with "{prefix}"'
                                                  + (' (case-insensitive)' if case_insensitive else ''),
                                    source=f'Legacy: {section_name}'))

        # AllowedValues: "Populate X using ... column"
        m = re.search(r'populate\s+(\w[\w\s]+?)\s+using', t)
        if m:
            field = _titlecase_field(m.group(1).strip())
            rules.append(_make_rule(field, 'AllowedValues', source=f'Legacy: {section_name}'))

    return rules


def _parse_reference_id(ws) -> dict:
    max_col = ws.max_column or 1
    max_row = ws.max_row or 1
    headers = [
        str(ws.cell(row=1, column=c).value or '').strip()
        for c in range(1, max_col + 1)
    ]
    allowed: dict = {}
    for col_idx, hdr in enumerate(headers, 1):
        if not hdr:
            continue
        vals = []
        for r in range(2, max_row + 1):
            v = str(ws.cell(row=r, column=col_idx).value or '').strip()
            if v:
                vals.append(v)
        if vals:
            allowed[hdr] = vals
    return allowed


# =======================================================================
# SECTION 2: SHEET DETECTOR  (from sheet_detector.py)
# =======================================================================

SKIP_SHEET_KEYWORDS = {
    'instructions', 'notes', 'legend', 'walkthrough',
    'examples', 'sample', 'cover', 'dq rules', 'dnu',
}

DATE_DATATYPE_TOKENS = {'dd-mmm-yyyy', 'dd/mm/yyyy', 'yyyy-mm-dd', 'date'}
NUMBER_DATATYPE_TOKENS = {'number', 'numeric', 'decimal', 'integer', 'int', 'float'}
BOOL_DATATYPE_TOKENS = {'true/false', 'yes/no', 'boolean', 'bool'}


def should_skip_sheet(sheet_name: str) -> bool:
    name_l = sheet_name.lower()
    return any(kw in name_l for kw in SKIP_SHEET_KEYWORDS) or name_l.startswith('>')


def detect_header_row(ws) -> tuple:
    """
    Returns (header_row_num, headers_list, datatype_map, req_optional_row_num).
    Returns (None, [], {}, None) if detection fails.
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row < 2 or max_col < 1:
        return None, [], {}, None

    req_row = _find_required_optional_row(ws, max_row, max_col)
    if req_row is None:
        return None, [], {}, None

    header_row = req_row + 1
    if header_row > max_row:
        return None, [], {}, None

    headers = [
        str(ws.cell(row=header_row, column=c).value or '').strip()
        for c in range(1, max_col + 1)
    ]

    datatype_row = req_row - 1
    datatype_map = {}
    if datatype_row >= 1:
        for col_idx, hdr in enumerate(headers, 1):
            if not hdr:
                continue
            dt_raw = str(ws.cell(row=datatype_row, column=col_idx).value or '').strip().lower()
            datatype_map[hdr] = _classify_datatype(dt_raw)

    return header_row, headers, datatype_map, req_row


def get_data_rows(ws, header_row: int, headers: list) -> list:
    """
    Returns list of dicts: {col_name: raw_value_str, '__row_num__': int, '__cells__': {col: cell}}.
    Skips fully blank rows.
    """
    max_row = ws.max_row or 0
    rows = []
    for r in range(header_row + 1, max_row + 1):
        row_dict = {'__row_num__': r, '__cells__': {}}
        is_blank = True
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=r, column=col_idx)
            val = cell.value
            val_str = str(val).strip() if val is not None else ''
            row_dict['__cells__'][hdr] = cell
            if hdr:
                row_dict[hdr] = val_str
                if val_str:
                    is_blank = False
        if not is_blank:
            rows.append(row_dict)
    return rows


def get_required_from_sheet(ws, req_row: int, headers: list) -> list:
    """Returns list of column names marked 'Required' in the req_row."""
    required = []
    max_col = ws.max_column or 0
    for col_idx, hdr in enumerate(headers, 1):
        if not hdr or col_idx > max_col:
            continue
        val = str(ws.cell(row=req_row, column=col_idx).value or '').strip().lower()
        if val == 'required':
            required.append(hdr)
    return required


def _find_required_optional_row(ws, max_row: int, max_col: int) -> Optional[int]:
    for r in range(1, min(max_row + 1, 20)):
        vals = [
            str(ws.cell(row=r, column=c).value or '').strip().lower()
            for c in range(1, max_col + 1)
        ]
        non_empty = [v for v in vals if v]
        if not non_empty:
            continue
        if len(non_empty) >= 1 and all(v in ('required', 'optional') for v in non_empty):
            return r
    return None


def _classify_datatype(dt_raw: str) -> str:
    if any(t in dt_raw for t in DATE_DATATYPE_TOKENS):
        return 'date'
    if any(t in dt_raw for t in NUMBER_DATATYPE_TOKENS):
        return 'number'
    if any(t in dt_raw for t in BOOL_DATATYPE_TOKENS):
        return 'boolean'
    if dt_raw.endswith('_id') or '_id' in dt_raw:
        return 'id'
    return 'text'


# =======================================================================
# SECTION 3: REPORT WRITER  (from report_writer.py)
# =======================================================================

HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
DATA_FONT = Font(name='Calibri', size=10)
TITLE_FONT = Font(name='Calibri', bold=True, size=13, color='FFFFFF')

PURPLE_FILL = PatternFill('solid', fgColor='6B21A8')
DARK_FILL   = PatternFill('solid', fgColor='4B5563')
RED_FILL    = PatternFill('solid', fgColor='FFB3B3')
ORANGE_FILL = PatternFill('solid', fgColor='FFE0B2')
YELLOW_FILL = PatternFill('solid', fgColor='FFF9C4')
ALT_FILL    = PatternFill('solid', fgColor='F9F9F9')
GREEN_FILL  = PatternFill('solid', fgColor='C8E6C9')
BLUE_FILL   = PatternFill('solid', fgColor='BBDEFB')

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center')

THIN_BORDER = Border(
    left=Side(style='thin',   color='D1D5DB'),
    right=Side(style='thin',  color='D1D5DB'),
    top=Side(style='thin',    color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)

SUMMARY_COLS = [
    'DGW Sheet Name', 'Matched Rule Section', 'Confidence Score',
    'Header Row', 'Total Columns', 'Total Data Rows', 'Total Checks',
    'Passed Checks', 'Failed Checks', 'Success %', 'Failed Rows', 'Passed Rows'
]

REASONS_COLS = [
    'Source Row Number', 'Primary Key', 'Failed Column', 'Invalid Value',
    'Validation Type', 'Rule Violated', 'Issue Description',
    'Allowed Values / Reference', 'Severity', 'Rule Source'
]


def _set_col_widths(ws: Worksheet, min_width: int = 10, max_width: int = 60):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _write_header_row(ws: Worksheet, row_num: int, headers: list, fill=None):
    fill = fill or PURPLE_FILL
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _severity_fill(severity: str) -> Optional[PatternFill]:
    s = severity.lower() if severity else ''
    if s == 'high':
        return RED_FILL
    if s == 'medium':
        return ORANGE_FILL
    if s == 'low':
        return YELLOW_FILL
    return None


def write_report(results: list, output_path: str, summary_stats: dict):
    wb = Workbook()

    idx_ws = wb.active
    idx_ws.title = 'Index'
    _write_index(idx_ws, results, summary_stats)

    for res in results:
        sheet_name = res.get('sheet_name', 'Unknown')
        safe_name = _safe_tab_name(sheet_name)

        summ_ws = wb.create_sheet(title=f'{safe_name[:25]}_Summ')
        _write_sheet_summary(summ_ws, res)

        rsn_ws = wb.create_sheet(title=f'{safe_name[:24]}_Rsns')
        _write_sheet_reasons(rsn_ws, res)

    wb.save(output_path)
    wb.close()


def _safe_tab_name(name: str) -> str:
    safe = re.sub(r'[\\/*?\[\]:]', '', name)
    return safe[:31]


def _write_index(ws: Worksheet, results: list, stats: dict):
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = 'DGW Validation Engine -- Validation Report'
    title_cell.font = TITLE_FONT
    title_cell.fill = PURPLE_FILL
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:L2')
    ts_cell = ws['A2']
    ts_cell.value = f'Generated: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ts_cell.font = Font(name='Calibri', size=10, italic=True, color='6B7280')
    ts_cell.alignment = CENTER
    ws.row_dimensions[2].height = 18

    stat_headers = ['Total Sheets', 'Total Rows', 'Total Checks', 'Passed', 'Failed', 'High', 'Medium', 'Low', 'Success %']
    stat_values = [
        stats.get('total_sheets', 0),
        stats.get('total_rows', 0),
        stats.get('total_checks', 0),
        stats.get('total_passed', 0),
        stats.get('total_failed', 0),
        stats.get('high', 0),
        stats.get('medium', 0),
        stats.get('low', 0),
        f"{stats.get('success_pct', 0):.1f}%",
    ]
    _write_header_row(ws, 4, stat_headers, DARK_FILL)
    for col_idx, v in enumerate(stat_values, 1):
        cell = ws.cell(row=5, column=col_idx, value=v)
        cell.font = DATA_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    _write_header_row(ws, 7, SUMMARY_COLS)
    for row_idx, res in enumerate(results, 8):
        is_alt = (row_idx % 2 == 0)
        sn = res.get('sheet_name', '')
        matched = res.get('matched_rule_section', '')
        confidence = res.get('confidence', 0.0)
        header_row_num = res.get('header_row', 1)
        total_cols = res.get('total_columns', 0)
        total_rows = res.get('total_rows', 0)
        total_checks = res.get('total_checks', 0)
        passed = res.get('passed_checks', 0)
        failed = res.get('failed_checks', 0)
        pct = (passed / total_checks * 100) if total_checks else 100.0
        failed_rows = res.get('failed_rows', 0)
        passed_rows = total_rows - failed_rows

        row_vals = [
            sn, matched, f'{confidence:.0%}', header_row_num,
            total_cols, total_rows, total_checks, passed, failed,
            f'{pct:.1f}%', failed_rows, passed_rows
        ]
        for col_idx, v in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font = DATA_FONT
            cell.fill = ALT_FILL if is_alt else PatternFill()
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        pct_cell = ws.cell(row=row_idx, column=10)
        if pct >= 95:
            pct_cell.fill = GREEN_FILL
        elif pct >= 75:
            pct_cell.fill = YELLOW_FILL
        else:
            pct_cell.fill = RED_FILL

    ws.freeze_panes = 'A8'
    ws.auto_filter.ref = f'A7:{get_column_letter(len(SUMMARY_COLS))}7'
    _set_col_widths(ws)


def _write_sheet_summary(ws: Worksheet, res: dict):
    sname = res.get('sheet_name', 'Unknown')

    ws.merge_cells('A1:L1')
    tc = ws['A1']
    tc.value = f'Sheet Summary -- {sname}'
    tc.font = TITLE_FONT
    tc.fill = PURPLE_FILL
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 28

    _write_header_row(ws, 3, SUMMARY_COLS)
    total_checks = res.get('total_checks', 0)
    passed = res.get('passed_checks', 0)
    failed = res.get('failed_checks', 0)
    pct = (passed / total_checks * 100) if total_checks else 100.0
    total_rows = res.get('total_rows', 0)
    failed_rows = res.get('failed_rows', 0)
    row_vals = [
        sname,
        res.get('matched_rule_section', ''),
        f'{res.get("confidence", 0):.0%}',
        res.get('header_row', 1),
        res.get('total_columns', 0),
        total_rows,
        total_checks,
        passed,
        failed,
        f'{pct:.1f}%',
        failed_rows,
        total_rows - failed_rows,
    ]
    for col_idx, v in enumerate(row_vals, 1):
        cell = ws.cell(row=4, column=col_idx, value=v)
        cell.font = DATA_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    ws.cell(row=6, column=1, value='COLUMN ISSUE SUMMARY').font = Font(name='Calibri', bold=True, size=11)
    col_issue_headers = ['Column Name', 'Total Checks', 'Passed', 'Failed', 'Failure Rate %', 'Top Issue']
    _write_header_row(ws, 7, col_issue_headers, DARK_FILL)

    col_stats = res.get('column_stats', {})
    for row_offset, (col_name, cstats) in enumerate(col_stats.items()):
        r = 8 + row_offset
        is_alt = row_offset % 2 == 0
        tc2 = cstats.get('total', 0)
        fc2 = cstats.get('failed', 0)
        pc2 = tc2 - fc2
        rate = (fc2 / tc2 * 100) if tc2 else 0.0
        top_issue = cstats.get('top_issue', '')
        row_data = [col_name, tc2, pc2, fc2, f'{rate:.1f}%', top_issue]
        for col_idx, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            cell.font = DATA_FONT
            cell.fill = ALT_FILL if is_alt else PatternFill()
            cell.alignment = LEFT if col_idx in (1, 6) else CENTER
            cell.border = THIN_BORDER

    offset_start = 8 + len(col_stats) + 2
    ws.cell(row=offset_start, column=1, value='VALIDATION CATEGORY SUMMARY').font = Font(name='Calibri', bold=True, size=11)
    cat_headers = ['Validation Type', 'Total Checks', 'Passed', 'Failed', 'Severity']
    _write_header_row(ws, offset_start + 1, cat_headers, DARK_FILL)

    cat_stats = res.get('category_stats', {})
    for row_offset, (cat_name, cat_data) in enumerate(cat_stats.items()):
        r = offset_start + 2 + row_offset
        is_alt = row_offset % 2 == 0
        tc3 = cat_data.get('total', 0)
        fc3 = cat_data.get('failed', 0)
        row_data = [cat_name, tc3, tc3 - fc3, fc3, cat_data.get('severity', '')]
        for col_idx, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            cell.font = DATA_FONT
            cell.fill = ALT_FILL if is_alt else PatternFill()
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    ws.freeze_panes = 'A3'
    _set_col_widths(ws)


def _write_sheet_reasons(ws: Worksheet, res: dict):
    sname = res.get('sheet_name', 'Unknown')

    ws.merge_cells(f'A1:{get_column_letter(len(REASONS_COLS))}1')
    tc = ws['A1']
    tc.value = f'Validation Failures -- {sname}'
    tc.font = TITLE_FONT
    tc.fill = PURPLE_FILL
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 28

    _write_header_row(ws, 2, REASONS_COLS)

    failures = res.get('failures', [])
    for row_offset, fail in enumerate(failures):
        r = 3 + row_offset
        is_alt = row_offset % 2 == 0
        severity = fail.get('severity', 'Low')
        sev_fill = _severity_fill(severity)

        row_data = [
            fail.get('row_num', ''),
            fail.get('primary_key', ''),
            fail.get('column', ''),
            fail.get('value', ''),
            fail.get('validation_type', ''),
            fail.get('rule', ''),
            fail.get('description', ''),
            fail.get('allowed_values', ''),
            severity,
            fail.get('rule_source', ''),
        ]
        for col_idx, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            cell.font = DATA_FONT
            if sev_fill:
                cell.fill = sev_fill
            elif is_alt:
                cell.fill = ALT_FILL
            cell.alignment = LEFT
            cell.border = THIN_BORDER

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(REASONS_COLS))}2'
    _set_col_widths(ws)


# =======================================================================
# SECTION 4: VALIDATION ENGINE  (from engine.py)
# =======================================================================

DATE_ISO_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')

def _parse_unique_cols_from_dq_rules(dgw_wb) -> Dict[str, List[str]]:
    """
    Dynamically read unique-column constraints from the '>Workday DQ Rules' sheet.
    Columns marked 'Yes' in the 'Unique (Yes)' column are treated as unique PKs.
    Returns {sheet_name_lower: [col1, col2, ...]}.
    """
    unique_map: Dict[str, List[str]] = {}
    for ws in dgw_wb.worksheets:
        if 'workday dq rules' not in ws.title.lower():
            continue
        max_col = ws.max_column or 1
        max_row = ws.max_row or 1
        # Find header row (row 1) and locate 'Worksheet', 'Column', 'Unique' cols
        headers = [str(ws.cell(row=1, column=c).value or '').strip().lower()
                   for c in range(1, max_col + 1)]
        ws_col  = next((i+1 for i, h in enumerate(headers) if 'worksheet' in h), None)
        col_col = next((i+1 for i, h in enumerate(headers) if h == 'column'), None)
        uniq_col= next((i+1 for i, h in enumerate(headers) if 'unique' in h), None)
        if not (ws_col and col_col and uniq_col):
            continue
        for r in range(2, max_row + 1):
            sheet_val  = str(ws.cell(row=r, column=ws_col).value  or '').strip()
            col_val    = str(ws.cell(row=r, column=col_col).value  or '').strip()
            unique_val = str(ws.cell(row=r, column=uniq_col).value or '').strip().lower()
            if sheet_val and col_val and unique_val == 'yes':
                key = sheet_val.lower()
                unique_map.setdefault(key, [])
                if col_val not in unique_map[key]:
                    unique_map[key].append(col_val)
        break   # only one DQ Rules sheet
    return unique_map


def _normalize_number_format(fmt: str) -> str:
    if not fmt:
        return ''
    nf = fmt.lower()
    nf = nf.replace('\\', '')
    nf = nf.split(';')[0].strip()
    nf = nf.rstrip('@').strip()
    return nf


def _display_date_value(cell) -> str:
    val = cell.value
    if isinstance(val, (datetime.datetime, datetime.date)):
        fmt = _normalize_number_format(str(cell.number_format or ''))
        if fmt == 'yyyy-mm-dd':
            return val.strftime('%Y-%m-%d')
        if isinstance(val, datetime.datetime):
            return val.strftime('%m/%d/%Y') if 'mm' in (cell.number_format or '').lower() else str(val.date())
        return str(val)
    return str(val) if val is not None else ''


def _is_valid_date_cell(cell) -> Tuple[bool, str]:
    val = cell.value
    fmt = _normalize_number_format(str(cell.number_format or ''))
    display = _display_date_value(cell)

    if isinstance(val, (datetime.datetime, datetime.date)):
        if fmt == 'yyyy-mm-dd':
            return True, val.strftime('%Y-%m-%d') if hasattr(val, 'strftime') else str(val)
        return False, display

    if isinstance(val, str):
        s = val.strip()
        if not s:
            return True, ''
        if DATE_ISO_RE.match(s):
            try:
                datetime.date.fromisoformat(s)
                return True, s
            except ValueError:
                return False, s
        return False, s

    if val is None or val == '':
        return True, ''

    return False, str(val)


def _parse_date(cell) -> Optional[datetime.date]:
    if cell is None:
        return None
    val = cell.value
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, str):
        s = val.strip()
        if DATE_ISO_RE.match(s):
            try:
                return datetime.date.fromisoformat(s)
            except ValueError:
                pass
    return None


def _normalize_col_name(name: str) -> str:
    return name.strip().lower()


def _find_col(headers: List[str], *candidates: str) -> Optional[str]:
    h_lower = {h.lower(): h for h in headers if h}
    for cand in candidates:
        if cand.lower() in h_lower:
            return h_lower[cand.lower()]
    return None


def _inc(stats: dict, key: str, subkey: str = ''):
    if key not in stats:
        stats[key] = {'total': 0, 'failed': 0}
    stats[key]['total'] += 1


def _fail(stats: dict, key: str):
    if key not in stats:
        stats[key] = {'total': 0, 'failed': 0}
    stats[key]['failed'] += 1


def _fail_cat(cat_stats: dict, cat: str):
    if cat not in cat_stats:
        cat_stats[cat] = {'total': 0, 'failed': 0}
    cat_stats[cat]['failed'] += 1


def _apply_rule(rule: dict, col: str, raw_str: str, cell, row: dict,
                seen_vals: dict, ref_allowed: dict, headers: list) -> Optional[dict]:
    """
    Apply a single rule to a single cell value.
    Returns a failure-info dict on violation, or None if the check passes.
    `seen_vals` is mutated in-place for Unique tracking.
    """
    rule_type     = rule['rule_type'].strip().lower()
    severity      = rule.get('severity', 'High') or 'High'
    custom_msg    = rule.get('error_message', '').strip()
    source        = rule.get('source', '')

    def _fail(desc: str, allowed: str = '', vtype: str = ''):
        return {
            'col': col, 'value': raw_str,
            'description': custom_msg or desc,
            'allowed': allowed,
            'severity': severity,
            'rule_source': source,
            'validation_type': vtype or rule['rule_type'],
            'rule': custom_msg or desc,
        }

    if rule_type == 'required':
        if not raw_str or raw_str == 'None':
            return _fail(f'Required field "{col}" is blank', '', 'Required Field')

    elif rule_type == 'unique':
        if col not in seen_vals:
            seen_vals[col] = set()
        if raw_str and raw_str != 'None':
            if raw_str in seen_vals[col]:
                return _fail(f'Duplicate value "{raw_str}" in "{col}" (must be unique)',
                             'Unique values only', 'Duplicate Check')
            seen_vals[col].add(raw_str)

    elif rule_type == 'date':
        if cell is not None:
            is_valid, display = _is_valid_date_cell(cell)
            if not is_valid:
                nf = str(cell.number_format or '')
                import datetime as _dt
                if isinstance(cell.value, (_dt.datetime, _dt.date)):
                    detail = f"Excel date formatted as '{nf}' — must be YYYY-MM-DD format"
                else:
                    detail = f"Text date '{display}' does not match YYYY-MM-DD format"
                return _fail(
                    f'{col} must follow YYYY-MM-DD. Value "{display}" with format "{nf}". {detail}',
                    'YYYY-MM-DD', 'Date Validation',
                )

    elif rule_type == 'allowedvalues':
        avs = ref_allowed.get(col)
        if avs:
            if raw_str.lower() not in {a.lower() for a in avs}:
                short_list = ', '.join(avs[:15]) + ('...' if len(avs) > 15 else '')
                return _fail(
                    f'Value "{raw_str}" is not in the allowed list for "{col}"',
                    ', '.join(avs), 'Allowed Values',
                )

    elif rule_type == 'allowedlist':
        avs = [v.strip() for v in rule.get('rule_value', '').split(',') if v.strip()]
        if avs and raw_str.lower() not in {a.lower() for a in avs}:
            return _fail(
                f'Value "{raw_str}" not in allowed list: {", ".join(avs)}',
                ', '.join(avs), 'Allowed List',
            )

    elif rule_type == 'regex':
        pattern = rule.get('rule_value', '')
        if pattern:
            try:
                if not re.match(pattern, raw_str):
                    return _fail(f'Value "{raw_str}" does not match pattern {pattern}',
                                 pattern, 'Regex')
            except re.error:
                pass  # Bad pattern — skip

    elif rule_type == 'maxlength':
        try:
            max_len = int(rule.get('rule_value', ''))
            if len(raw_str) > max_len:
                return _fail(f'Value length {len(raw_str)} exceeds maximum {max_len}',
                             f'<= {max_len} chars', 'Max Length')
        except (ValueError, TypeError):
            pass

    elif rule_type == 'minlength':
        try:
            min_len = int(rule.get('rule_value', ''))
            if len(raw_str) < min_len:
                return _fail(f'Value length {len(raw_str)} is below minimum {min_len}',
                             f'>= {min_len} chars', 'Min Length')
        except (ValueError, TypeError):
            pass

    elif rule_type == 'numeric':
        try:
            float(raw_str.replace(',', ''))
        except (ValueError, AttributeError):
            return _fail(f'Value "{raw_str}" is not a valid number', 'Numeric', 'Numeric')

    elif rule_type == 'boolean':
        if raw_str.lower() not in {'yes', 'no', 'true', 'false', 'y', 'n', '1', '0', 'x'}:
            return _fail(f'Value "{raw_str}" is not a valid boolean',
                         'Yes/No/True/False/Y/N/1/0', 'Boolean')

    elif rule_type == 'daterange':
        other_col = rule.get('rule_value', '').strip()
        if other_col and other_col in row:
            other_cell = row.get('__cells__', {}).get(other_col)
            this_date  = _parse_date(cell)
            other_date = _parse_date(other_cell) if other_cell else None
            if this_date and other_date and this_date < other_date:
                return _fail(
                    f'{col} ({this_date}) cannot be before {other_col} ({other_date})',
                    f'>= {other_date}', 'Date Range',
                )

    elif rule_type == 'conditionalrequired':
        cond_field = rule.get('condition_field', '').strip()
        cond_val   = rule.get('condition_value', '').strip()
        if cond_field and cond_val:
            trigger = row.get(cond_field, '').strip()
            if trigger.lower() == cond_val.lower() and (not raw_str or raw_str == 'None'):
                return _fail(
                    f'"{col}" is required when "{cond_field}" = "{cond_val}"',
                    '', 'Conditional Required',
                )

    return None  # Rule passed


class ValidationEngine:

    def __init__(self, rules_path: str, dgw_path: str, progress_callback: Callable = None):
        self.rules_path = rules_path
        self.dgw_path = dgw_path
        self._progress = progress_callback or (lambda pct, msg: None)

    def run(self, output_path: str) -> dict:
        self._progress(10, 'Loading workbooks...')
        dgw_wb = load_workbook(self.dgw_path, data_only=True)

        self._progress(20, 'Discovering validation rules...')
        # Fully dynamic -- discovers every sheet in the rules workbook
        rules_cache = parse_rules_workbook(self.rules_path)

        self._progress(25, 'Reading DQ uniqueness rules...')
        # Fully dynamic -- reads Unique constraints from >Workday DQ Rules sheet
        unique_map = _parse_unique_cols_from_dq_rules(dgw_wb)

        self._progress(30, 'Detecting business sheets...')
        business_sheets = [
            ws for ws in dgw_wb.worksheets
            if not should_skip_sheet(ws.title)
        ]

        results = []
        total = len(business_sheets)

        hired_employee_ids: Set[str] = set()
        hired_position_ids: Set[str] = set()

        for idx, ws in enumerate(business_sheets):
            pct = 40 + int((idx / max(total, 1)) * 45)
            self._progress(pct, f'Validating {ws.title}...')
            result = self._validate_sheet(ws, rules_cache, unique_map,
                                          hired_employee_ids, hired_position_ids)
            results.append(result)

            sn = ws.title.lower()
            if sn == 'hire employee':
                for row in result.get('_rows', []):
                    eid = row.get('Employee ID', '').strip()
                    pid = row.get('Position ID', '').strip()
                    if eid:
                        hired_employee_ids.add(eid)
                    if pid:
                        hired_position_ids.add(pid)

        self._progress(90, 'Generating validation report...')
        summary = self._build_summary(results)
        write_report(results, output_path, summary)
        dgw_wb.close()
        self._progress(100, 'Complete')
        return summary

    def _validate_sheet(self, ws, rules_cache: dict, unique_map: dict,
                        hired_ids: Set[str], hired_pos_ids: Set[str]) -> dict:
        sname   = ws.title
        sname_l = sname.lower()

        header_row, headers, datatype_map, req_row = detect_header_row(ws)
        if header_row is None:
            return self._empty_result(sname)

        rows = get_data_rows(ws, header_row, headers)
        if not rows:
            return self._empty_result(sname)

        # Dynamic match: finds best rules entry for this sheet name
        sheet_rules_entry = _match_rules_for_sheet(sname, rules_cache)
        rules_list        = list(sheet_rules_entry.get('rules', []))
        ref_allowed       = sheet_rules_entry.get('allowed_values', {})
        matched_section   = sheet_rules_entry.get('matched_section', '')

        headers_set = set(h for h in headers if h)

        # Determine primary key column
        req_fields_from_rules = [r['field'] for r in rules_list
                                  if r['rule_type'].lower() == 'required']
        if not req_fields_from_rules and req_row is not None:
            req_fields_from_rules = get_required_from_sheet(ws, req_row, headers)

        pk_col = req_fields_from_rules[0] if req_fields_from_rules else (headers[0] if headers else '')
        for candidate in ['Employee ID', 'Contingent Worker ID', 'Worker ID', 'Associate ID']:
            if candidate in headers_set:
                pk_col = candidate
                break

        # --- Supplement rules_list from external sources ---

        existing_required = {r['field'] for r in rules_list if r['rule_type'].lower() == 'required'}

        # 1. Required rules from DGW row-5 metadata (for every field marked Required)
        #    Applied regardless of whether a rules sheet matched — ensures sheets with
        #    no matching rules entry (e.g. Add Additional Job) still get required checks.
        if req_row is not None:
            for col in get_required_from_sheet(ws, req_row, headers):
                if col not in existing_required:
                    rules_list.append(_make_rule(col, 'Required',
                                                 source='DGW Row 5 Metadata'))
                    existing_required.add(col)

        # 2. Unique rules from >Workday DQ Rules sheet
        for ucol in unique_map.get(sname_l, []):
            if not any(r['field'] == ucol and r['rule_type'].lower() == 'unique'
                       for r in rules_list):
                rules_list.append(_make_rule(ucol, 'Unique',
                                             source='>Workday DQ Rules'))

        # 3. Date rules from DGW row-4 metadata for columns not already covered
        for col, dtype in datatype_map.items():
            if dtype == 'date':
                if not any(r['field'] == col and r['rule_type'].lower() == 'date'
                           for r in rules_list):
                    rules_list.append(_make_rule(col, 'Date',
                                                 source='DGW Row 4 Metadata'))

        # --- Apply rules engine ---
        failures:      List[dict]          = []
        total_checks   = 0
        failed_checks  = 0
        failed_row_set: Set[int]           = set()
        column_stats:   Dict[str, dict]    = {}
        category_stats: Dict[str, dict]   = {}
        seen_vals:      Dict[str, Set[str]] = {}

        for row in rows:
            row_num = row['__row_num__']
            pk_val  = row.get(pk_col, '') if pk_col else ''

            for rule in rules_list:
                col = rule['field']
                if col not in headers_set:
                    continue

                raw_str  = row.get(col, '')
                cell     = row['__cells__'].get(col)
                is_blank = not raw_str or raw_str == 'None'

                # Non-required rules are skipped for blank cells
                # (Required and ConditionalRequired must always be evaluated)
                if is_blank and rule['rule_type'].lower() not in ('required', 'conditionalrequired'):
                    # For Unique we still need to note the cell was blank (skip tracking)
                    continue

                total_checks += 1
                _inc(column_stats, col)
                vtype = rule['rule_type']
                _inc(category_stats, vtype)

                failure = _apply_rule(rule, col, raw_str, cell, row,
                                      seen_vals, ref_allowed, headers)
                if failure:
                    failed_checks += 1
                    failed_row_set.add(row_num)
                    _fail(column_stats, col)
                    _fail_cat(category_stats, vtype)
                    failures.append(self._f(
                        row_num, pk_val,
                        failure.get('col', col),
                        failure.get('value', raw_str),
                        failure.get('validation_type', vtype),
                        failure.get('rule', ''),
                        failure.get('description', ''),
                        failure.get('allowed', ''),
                        failure.get('severity', 'High'),
                        failure.get('rule_source', matched_section or 'Rules Engine'),
                    ))

        # Legacy hardcoded date-range checks (award sheets) — kept for backward compat
        dr_failures = self._check_date_ranges(ws, headers, rows, sname_l, pk_col, matched_section)
        for f in dr_failures:
            failures.append(f)
            failed_checks += 1
            total_checks  += 1
            failed_row_set.add(f['row_num'])
            _inc(category_stats, 'Date Range')
            _fail_cat(category_stats, 'Date Range')

        passed_checks = total_checks - failed_checks

        return {
            'sheet_name':           sname,
            'matched_rule_section': matched_section,
            'confidence':           0.95 if matched_section else 0.70,
            'header_row':           header_row,
            'total_columns':        len([h for h in headers if h]),
            'total_rows':           len(rows),
            'total_checks':         total_checks,
            'passed_checks':        passed_checks,
            'failed_checks':        failed_checks,
            'failed_rows':          len(failed_row_set),
            'failures':             failures,
            'column_stats':         column_stats,
            'category_stats':       category_stats,
            '_rows':                rows,
            '_headers':             headers,
        }

    def _check_date_ranges(self, ws, headers, rows, sname_l, pk_col, rule_source):
        failures = []
        if 'award' not in sname_l:
            return failures

        if 'award line' in sname_l and 'worktag' not in sname_l:
            start_col = _find_col(headers, 'Award Line Start Date', 'Start Date')
            end_col   = _find_col(headers, 'Award Line End Date', 'End Date')
            for row in rows:
                rn = row['__row_num__']
                pk = row.get(pk_col, '') if pk_col else ''
                if not start_col or not end_col:
                    break
                sc = row['__cells__'].get(start_col)
                ec = row['__cells__'].get(end_col)
                if not sc or not ec:
                    continue
                sd, ed = _parse_date(sc), _parse_date(ec)
                if sd and ed:
                    if ed < sd:
                        failures.append(self._f(rn, pk, end_col, str(ed),
                            'Date Range', 'End Date >= Start Date',
                            f'Award Line End Date ({ed}) must be >= Start Date ({sd})',
                            f'>= {sd}', 'High', rule_source))
                    elif (ed - sd).days != 365:
                        failures.append(self._f(rn, pk, end_col, str(ed),
                            'Date Range', 'Award Line duration must be exactly 365 days',
                            f'Duration is {(ed-sd).days} days (expected 365)',
                            f'{sd} + 365 days', 'High', rule_source))

        elif 'award task' in sname_l or ('task' in sname_l and 'award' in sname_l):
            start_col = _find_col(headers, 'Start Date', 'Task Start Date')
            end_col   = _find_col(headers, 'End Date', 'Task End Date')
            for row in rows:
                rn = row['__row_num__']
                pk = row.get(pk_col, '') if pk_col else ''
                if not start_col or not end_col:
                    break
                sd = _parse_date(row['__cells__'].get(start_col))
                ed = _parse_date(row['__cells__'].get(end_col))
                if sd and ed and ed < sd:
                    failures.append(self._f(rn, pk, end_col, str(ed),
                        'Date Range', 'End Date >= Start Date',
                        f'Task End Date ({ed}) must be >= Start Date ({sd})',
                        f'>= {sd}', 'High', rule_source))

        elif 'award' in sname_l and 'line' not in sname_l and 'task' not in sname_l:
            signed_col    = _find_col(headers, 'Award Signed Date')
            effective_col = _find_col(headers, 'Award Effective Date')
            for row in rows:
                rn = row['__row_num__']
                pk = row.get(pk_col, '') if pk_col else ''
                if not signed_col or not effective_col:
                    break
                sd = _parse_date(row['__cells__'].get(signed_col))
                ed = _parse_date(row['__cells__'].get(effective_col))
                if sd and ed and sd < ed:
                    failures.append(self._f(rn, pk, signed_col, str(sd),
                        'Date Range', 'Award Signed Date >= Award Effective Date',
                        f'Signed Date ({sd}) must be >= Effective Date ({ed})',
                        f'>= {ed}', 'High', rule_source))

        return failures

    @staticmethod
    def _f(row_num, pk, col, val, vtype, rule, desc, allowed, severity, source) -> dict:
        return {
            'row_num': row_num, 'primary_key': pk,
            'column': col, 'value': val,
            'validation_type': vtype, 'rule': rule,
            'description': desc, 'allowed_values': allowed,
            'severity': severity, 'rule_source': source,
        }

    @staticmethod
    def _empty_result(sname) -> dict:
        return {
            'sheet_name': sname, 'matched_rule_section': '', 'confidence': 0.0,
            'header_row': None, 'total_columns': 0, 'total_rows': 0,
            'total_checks': 0, 'passed_checks': 0, 'failed_checks': 0,
            'failed_rows': 0, 'failures': [], 'column_stats': {}, 'category_stats': {},
            '_rows': [], '_headers': [],
        }

    def _build_summary(self, results):
        total_sheets  = len(results)
        total_rows    = sum(r.get('total_rows', 0) for r in results)
        total_checks  = sum(r.get('total_checks', 0) for r in results)
        total_passed  = sum(r.get('passed_checks', 0) for r in results)
        total_failed  = sum(r.get('failed_checks', 0) for r in results)
        high   = sum(sum(1 for f in r.get('failures', []) if f.get('severity') == 'High')   for r in results)
        medium = sum(sum(1 for f in r.get('failures', []) if f.get('severity') == 'Medium') for r in results)
        low    = sum(sum(1 for f in r.get('failures', []) if f.get('severity') == 'Low')    for r in results)
        success_pct = (total_passed / total_checks * 100) if total_checks else 100.0
        return {
            'total_sheets': total_sheets, 'total_rows': total_rows,
            'total_checks': total_checks, 'total_passed': total_passed,
            'total_failed': total_failed, 'high': high, 'medium': medium,
            'low': low, 'success_pct': success_pct,
        }


# =======================================================================
# SECTION 5: HTML / CSS / JS TEMPLATES  (embedded strings)
# =======================================================================

CSS_CONTENT = """
/* =========================================================
   DGW Validation Engine -- Stylesheet
   Accenture-inspired: purple accent #A100FF
   ========================================================= */

:root {
  --accent:        #A100FF;
  --accent-dark:   #7B00CC;
  --accent-light:  #E8B3FF;
  --accent-bg:     #F5E6FF;
  --bg:            #F3F4F6;
  --card-bg:       #FFFFFF;
  --border:        #E5E7EB;
  --text-primary:  #111827;
  --text-secondary:#6B7280;
  --text-muted:    #9CA3AF;
  --success:       #10B981;
  --success-bg:    #D1FAE5;
  --warning:       #F59E0B;
  --warning-bg:    #FEF3C7;
  --danger:        #EF4444;
  --danger-bg:     #FEE2E2;
  --info:          #3B82F6;
  --info-bg:       #DBEAFE;
  --shadow-sm:     0 1px 3px rgba(0,0,0,0.08);
  --shadow-md:     0 4px 12px rgba(0,0,0,0.10);
  --shadow-lg:     0 8px 24px rgba(0,0,0,0.12);
  --radius:        12px;
  --radius-sm:     8px;
  --transition:    0.2s ease;
}

*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

body {
  font-family: 'Inter', 'Segoe UI', sans-serif;
  background: var(--bg);
  color: var(--text-primary);
  min-height: 100vh;
  line-height: 1.5;
}

/* -- Header -- */
.app-header {
  background: linear-gradient(135deg, #1a0033 0%, #4B0082 60%, #A100FF 100%);
  color: #fff;
  padding: 28px 0 24px;
  text-align: center;
  position: relative;
  overflow: hidden;
}
.app-header::before {
  content: '';
  position: absolute;
  inset: 0;
  background: url("data:image/svg+xml,%3Csvg width='60' height='60' viewBox='0 0 60 60' xmlns='http://www.w3.org/2000/svg'%3E%3Cg fill='none' fill-rule='evenodd'%3E%3Cg fill='%23ffffff' fill-opacity='0.03'%3E%3Cpath d='M36 34v-4h-2v4h-4v2h4v4h2v-4h4v-2h-4zm0-30V0h-2v4h-4v2h4v4h2V6h4V4h-4zM6 34v-4H4v4H0v2h4v4h2v-4h4v-2H6zM6 4V0H4v4H0v2h4v4h2V6h4V4H6z'/%3E%3C/g%3E%3C/g%3E%3C/svg%3E");
}
.app-header .inner { position: relative; }
.app-header h1 {
  font-size: 2rem;
  font-weight: 700;
  letter-spacing: -0.5px;
  margin-bottom: 4px;
}
.app-header .subtitle {
  font-size: 0.95rem;
  opacity: 0.8;
  margin-bottom: 10px;
}
.version-badge {
  display: inline-block;
  background: rgba(255,255,255,0.15);
  border: 1px solid rgba(255,255,255,0.3);
  color: #fff;
  font-size: 0.72rem;
  font-weight: 600;
  padding: 3px 10px;
  border-radius: 999px;
  letter-spacing: 0.5px;
  backdrop-filter: blur(4px);
}

/* -- Main container -- */
.container {
  max-width: 1100px;
  margin: 0 auto;
  padding: 32px 20px 60px;
}

/* -- Cards -- */
.card {
  background: var(--card-bg);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  box-shadow: var(--shadow-sm);
  padding: 28px;
  margin-bottom: 24px;
  transition: box-shadow var(--transition);
}
.card:hover { box-shadow: var(--shadow-md); }
.card-title {
  font-size: 1.05rem;
  font-weight: 600;
  color: var(--text-primary);
  margin-bottom: 20px;
  display: flex;
  align-items: center;
  gap: 8px;
}
.card-title .icon {
  width: 28px; height: 28px;
  background: var(--accent-bg);
  border-radius: 7px;
  display: flex; align-items: center; justify-content: center;
  font-size: 14px;
  flex-shrink: 0;
}

/* -- Upload grid -- */
.upload-grid {
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 20px;
  margin-bottom: 24px;
}
@media (max-width: 640px) { .upload-grid { grid-template-columns: 1fr; } }

/* -- Drop zone -- */
.drop-zone {
  border: 2px dashed var(--border);
  border-radius: var(--radius-sm);
  padding: 36px 20px;
  text-align: center;
  cursor: pointer;
  transition: border-color var(--transition), background var(--transition);
  position: relative;
  background: #FAFAFA;
}
.drop-zone:hover,
.drop-zone.drag-over {
  border-color: var(--accent);
  background: var(--accent-bg);
}
.drop-zone.file-selected {
  border-color: var(--success);
  background: var(--success-bg);
}
.drop-zone input[type="file"] {
  position: absolute;
  inset: 0;
  opacity: 0;
  cursor: pointer;
  width: 100%;
  height: 100%;
}
.drop-zone .dz-icon {
  font-size: 2.2rem;
  margin-bottom: 10px;
  display: block;
}
.drop-zone .dz-label {
  font-size: 0.875rem;
  font-weight: 600;
  color: var(--text-primary);
  margin-bottom: 4px;
}
.drop-zone .dz-hint {
  font-size: 0.78rem;
  color: var(--text-muted);
}
.drop-zone .dz-filename {
  margin-top: 12px;
  font-size: 0.82rem;
  color: var(--success);
  font-weight: 600;
  display: flex;
  align-items: center;
  justify-content: center;
  gap: 6px;
}
.checkmark { color: var(--success); font-size: 1rem; }

/* -- Run button -- */
.btn-run {
  display: block;
  width: 100%;
  padding: 14px;
  background: var(--accent);
  color: #fff;
  font-size: 1rem;
  font-weight: 700;
  border: none;
  border-radius: var(--radius-sm);
  cursor: pointer;
  transition: background var(--transition), transform var(--transition), box-shadow var(--transition);
  letter-spacing: 0.3px;
}
.btn-run:hover:not(:disabled) {
  background: var(--accent-dark);
  transform: translateY(-1px);
  box-shadow: 0 6px 20px rgba(161,0,255,0.35);
}
.btn-run:active:not(:disabled) { transform: translateY(0); }
.btn-run:disabled {
  background: #D1D5DB;
  color: #9CA3AF;
  cursor: not-allowed;
  box-shadow: none;
}

/* -- Progress section -- */
#progress-section {
  display: none;
  animation: fadeIn 0.3s ease;
}
#progress-section.visible { display: block; }

.progress-track {
  height: 8px;
  background: var(--border);
  border-radius: 999px;
  overflow: hidden;
  margin-bottom: 16px;
}
.progress-bar {
  height: 100%;
  background: linear-gradient(90deg, var(--accent-dark), var(--accent));
  border-radius: 999px;
  transition: width 0.4s ease;
  position: relative;
  overflow: hidden;
}
.progress-bar::after {
  content: '';
  position: absolute;
  top: 0; left: -100%;
  width: 100%; height: 100%;
  background: linear-gradient(90deg, transparent, rgba(255,255,255,0.4), transparent);
  animation: shimmer 1.5s infinite;
}
@keyframes shimmer {
  0%   { left: -100%; }
  100% { left: 200%; }
}

.progress-pct {
  font-size: 0.95rem;
  font-weight: 700;
  color: var(--accent);
  margin-bottom: 8px;
  text-align: right;
}
.progress-current {
  font-size: 0.9rem;
  color: var(--text-secondary);
  margin-bottom: 16px;
  display: flex;
  align-items: center;
  gap: 8px;
}
.spinner {
  width: 16px; height: 16px;
  border: 2px solid var(--accent-light);
  border-top-color: var(--accent);
  border-radius: 50%;
  animation: spin 0.8s linear infinite;
  flex-shrink: 0;
}
@keyframes spin { to { transform: rotate(360deg); } }

.step-log {
  background: #F9FAFB;
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  padding: 12px 16px;
  max-height: 160px;
  overflow-y: auto;
  font-size: 0.8rem;
  color: var(--text-secondary);
}
.step-log p {
  padding: 3px 0;
  border-bottom: 1px solid var(--border);
  display: flex;
  align-items: center;
  gap: 8px;
}
.step-log p:last-child { border-bottom: none; }
.step-log p::before { content: '\\2713'; color: var(--success); font-weight: 700; }

/* -- Results section -- */
#results-section {
  display: none;
  animation: fadeIn 0.4s ease;
}
#results-section.visible { display: block; }

/* -- Summary cards -- */
.summary-grid {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
  gap: 16px;
  margin-bottom: 24px;
}
.stat-card {
  background: var(--card-bg);
  border: 1px solid var(--border);
  border-radius: var(--radius-sm);
  padding: 20px 16px;
  text-align: center;
  box-shadow: var(--shadow-sm);
  transition: transform var(--transition);
}
.stat-card:hover { transform: translateY(-2px); }
.stat-card .stat-value {
  font-size: 2rem;
  font-weight: 800;
  line-height: 1;
  margin-bottom: 6px;
}
.stat-card .stat-label {
  font-size: 0.75rem;
  color: var(--text-secondary);
  font-weight: 500;
  text-transform: uppercase;
  letter-spacing: 0.5px;
}
.stat-card.accent .stat-value { color: var(--accent); }
.stat-card.success .stat-value { color: var(--success); }
.stat-card.warning .stat-value { color: var(--warning); }
.stat-card.danger  .stat-value { color: var(--danger); }
.stat-card.info    .stat-value { color: var(--info); }

/* -- Per-sheet table -- */
.table-wrapper {
  overflow-x: auto;
  border-radius: var(--radius-sm);
  border: 1px solid var(--border);
}
table {
  width: 100%;
  border-collapse: collapse;
  font-size: 0.85rem;
}
thead tr {
  background: linear-gradient(135deg, #4B0082, var(--accent));
  color: #fff;
}
thead th {
  padding: 12px 14px;
  text-align: left;
  font-weight: 600;
  white-space: nowrap;
  font-size: 0.78rem;
  letter-spacing: 0.3px;
}
tbody tr {
  border-bottom: 1px solid var(--border);
  transition: background var(--transition);
}
tbody tr:last-child { border-bottom: none; }
tbody tr:hover { background: var(--accent-bg); }
tbody tr:nth-child(even) { background: #FAFAFA; }
tbody tr:nth-child(even):hover { background: var(--accent-bg); }
tbody td {
  padding: 10px 14px;
  vertical-align: middle;
}
.badge {
  display: inline-block;
  padding: 2px 8px;
  border-radius: 999px;
  font-size: 0.72rem;
  font-weight: 600;
}
.badge-high   { background: var(--danger-bg);  color: #B91C1C; }
.badge-medium { background: var(--warning-bg); color: #92400E; }
.badge-low    { background: var(--info-bg);    color: #1D4ED8; }
.pct-bar-wrap {
  display: flex;
  align-items: center;
  gap: 8px;
}
.pct-bar-track {
  flex: 1;
  height: 6px;
  background: var(--border);
  border-radius: 999px;
  overflow: hidden;
}
.pct-bar-fill {
  height: 100%;
  border-radius: 999px;
  transition: width 0.5s ease;
}

/* -- Download button -- */
.btn-download {
  display: inline-flex;
  align-items: center;
  gap: 10px;
  padding: 14px 32px;
  background: var(--success);
  color: #fff;
  font-size: 1rem;
  font-weight: 700;
  border: none;
  border-radius: var(--radius-sm);
  cursor: pointer;
  text-decoration: none;
  transition: background var(--transition), transform var(--transition), box-shadow var(--transition);
  margin-top: 24px;
}
.btn-download:hover {
  background: #059669;
  transform: translateY(-1px);
  box-shadow: 0 6px 20px rgba(16,185,129,0.4);
}
.btn-download:active { transform: translateY(0); }

/* -- Error banner -- */
.error-banner {
  background: var(--danger-bg);
  border: 1px solid #FECACA;
  border-radius: var(--radius-sm);
  padding: 16px 20px;
  color: #B91C1C;
  display: none;
  margin-bottom: 16px;
  font-size: 0.875rem;
  animation: fadeIn 0.3s ease;
}
.error-banner.visible { display: flex; align-items: flex-start; gap: 10px; }

/* -- Animations -- */
@keyframes fadeIn {
  from { opacity: 0; transform: translateY(8px); }
  to   { opacity: 1; transform: translateY(0); }
}

/* -- Scrollbar styling -- */
.step-log::-webkit-scrollbar { width: 5px; }
.step-log::-webkit-scrollbar-track { background: var(--bg); }
.step-log::-webkit-scrollbar-thumb { background: var(--border); border-radius: 999px; }
"""

JS_CONTENT = """
/* =========================================================
   DGW Validation Engine -- Frontend JS
   ========================================================= */

'use strict';

// State
let rulesFile = null;
let dgwFile = null;
let currentJobId = null;
let pollInterval = null;

// DOM refs (populated on DOMContentLoaded)
let elRulesZone, elDgwZone, elRunBtn, elProgressSection, elResultsSection,
    elErrorBanner, elProgressBar, elProgressPct, elProgressCurrent, elStepLog;

// ---------------------------------------------------------------------------
// Init
// ---------------------------------------------------------------------------
document.addEventListener('DOMContentLoaded', () => {
  elRulesZone       = document.getElementById('rules-zone');
  elDgwZone         = document.getElementById('dgw-zone');
  elRunBtn          = document.getElementById('run-btn');
  elProgressSection = document.getElementById('progress-section');
  elResultsSection  = document.getElementById('results-section');
  elErrorBanner     = document.getElementById('error-banner');
  elProgressBar     = document.getElementById('progress-bar');
  elProgressPct     = document.getElementById('progress-pct');
  elProgressCurrent = document.getElementById('progress-current');
  elStepLog         = document.getElementById('step-log');

  initDropZone(elRulesZone, 'rules');
  initDropZone(elDgwZone, 'dgw');

  elRunBtn.addEventListener('click', runValidation);
});

// ---------------------------------------------------------------------------
// Drop zone
// ---------------------------------------------------------------------------
function initDropZone(zone, type) {
  const input = zone.querySelector('input[type="file"]');
  const fnDisplay = zone.querySelector('.dz-filename');

  input.addEventListener('change', () => {
    const file = input.files[0];
    if (file) setFile(type, file, zone, fnDisplay);
  });

  zone.addEventListener('dragover', e => { e.preventDefault(); zone.classList.add('drag-over'); });
  zone.addEventListener('dragleave', () => zone.classList.remove('drag-over'));
  zone.addEventListener('drop', e => {
    e.preventDefault();
    zone.classList.remove('drag-over');
    const file = e.dataTransfer.files[0];
    if (file) {
      if (!file.name.endsWith('.xlsx')) {
        showError('Only .xlsx files are accepted.');
        return;
      }
      setFile(type, file, zone, fnDisplay);
    }
  });
}

function setFile(type, file, zone, fnDisplay) {
  if (!file.name.endsWith('.xlsx')) {
    showError('Only .xlsx files are accepted.');
    return;
  }
  if (type === 'rules') rulesFile = file;
  else dgwFile = file;

  zone.classList.add('file-selected');
  fnDisplay.innerHTML = '<span class="checkmark">\\u2713</span> ' + escHtml(file.name);
  fnDisplay.style.display = 'flex';

  updateRunButton();
}

function updateRunButton() {
  elRunBtn.disabled = !(rulesFile && dgwFile);
}

// ---------------------------------------------------------------------------
// Validation submission
// ---------------------------------------------------------------------------
async function runValidation() {
  hideError();
  resetResults();

  const formData = new FormData();
  formData.append('rules_file', rulesFile);
  formData.append('dgw_file', dgwFile);

  showProgress();

  try {
    const resp = await fetch('/validate', { method: 'POST', body: formData });
    if (!resp.ok) {
      const err = await resp.json();
      throw new Error(err.error || 'Upload failed');
    }
    const data = await resp.json();
    currentJobId = data.job_id;
    startPolling();
  } catch (err) {
    showError(err.message || 'Failed to start validation.');
    hideProgress();
  }
}

// ---------------------------------------------------------------------------
// Polling
// ---------------------------------------------------------------------------
function startPolling() {
  if (pollInterval) clearInterval(pollInterval);
  pollInterval = setInterval(poll, 1500);
}

async function poll() {
  if (!currentJobId) return;
  try {
    const resp = await fetch('/status/' + currentJobId);
    if (!resp.ok) return;
    const data = await resp.json();
    updateProgress(data);

    if (data.status === 'complete') {
      clearInterval(pollInterval);
      pollInterval = null;
      renderResults(data.summary);
    } else if (data.status === 'error') {
      clearInterval(pollInterval);
      pollInterval = null;
      showError(data.error || 'Validation failed.');
      hideProgress();
    }
  } catch (_) {
    // network hiccup -- keep polling
  }
}

// ---------------------------------------------------------------------------
// Progress UI
// ---------------------------------------------------------------------------
function showProgress() {
  elProgressSection.classList.add('visible');
  elResultsSection.classList.remove('visible');
  setProgress(0, 'Starting...');
}

function hideProgress() {
  elProgressSection.classList.remove('visible');
}

function updateProgress(data) {
  setProgress(data.progress || 0, data.current || '');
  const log = elStepLog;
  const existing = Array.from(log.querySelectorAll('p')).map(p => p.dataset.step);
  (data.steps || []).forEach(step => {
    if (!existing.includes(step)) {
      const p = document.createElement('p');
      p.dataset.step = step;
      p.appendChild(document.createTextNode(step));
      log.appendChild(p);
    }
  });
  log.scrollTop = log.scrollHeight;
}

function setProgress(pct, msg) {
  elProgressBar.style.width = pct + '%';
  elProgressPct.textContent = pct + '%';
  const span = elProgressCurrent.querySelector('span');
  if (span) span.textContent = msg;
}

// ---------------------------------------------------------------------------
// Results
// ---------------------------------------------------------------------------
function renderResults(summary) {
  if (!summary) return;

  document.getElementById('stat-sheets').textContent  = fmt(summary.total_sheets);
  document.getElementById('stat-rows').textContent    = fmt(summary.total_rows);
  document.getElementById('stat-issues').textContent  = fmt(summary.total_failed);
  document.getElementById('stat-high').textContent    = fmt(summary.high);
  document.getElementById('stat-medium').textContent  = fmt(summary.medium);
  document.getElementById('stat-low').textContent     = fmt(summary.low);
  document.getElementById('stat-success').textContent = (summary.success_pct || 0).toFixed(1) + '%';

  const dlBtn = document.getElementById('download-btn');
  dlBtn.href = '/download/' + currentJobId;
  dlBtn.style.display = 'inline-flex';

  elResultsSection.classList.add('visible');
}

function resetResults() {
  elStepLog.innerHTML = '';
  elResultsSection.classList.remove('visible');
  document.getElementById('download-btn').style.display = 'none';
}

// ---------------------------------------------------------------------------
// Error
// ---------------------------------------------------------------------------
function showError(msg) {
  elErrorBanner.classList.add('visible');
  const spans = elErrorBanner.querySelectorAll('span');
  if (spans.length > 1) spans[1].textContent = msg;
}
function hideError() {
  elErrorBanner.classList.remove('visible');
}

// ---------------------------------------------------------------------------
// Utilities
// ---------------------------------------------------------------------------
function escHtml(str) {
  return str.replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;').replace(/"/g, '&quot;');
}
function fmt(n) {
  if (n === undefined || n === null) return '\\u2014';
  return Number(n).toLocaleString();
}
"""

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>DGW Validation Engine</title>
  <link rel="preconnect" href="https://fonts.googleapis.com" />
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet" />
  <link rel="stylesheet" href="/static/style.css" />
</head>
<body>

<!-- ======================================================
     Header
======================================================== -->
<header class="app-header">
  <div class="inner">
    <h1>&#x26A1; DGW Validation Engine</h1>
    <p class="subtitle">Workday Data Quality Validator</p>
    <span class="version-badge">v3.0</span>
  </div>
</header>

<!-- ======================================================
     Main content
======================================================== -->
<main class="container">

  <!-- Error banner -->
  <div class="error-banner" id="error-banner" role="alert">
    <span style="font-size:1.2rem;">&#x26A0;&#xFE0F;</span>
    <span>Error message here</span>
  </div>

  <!-- -- Upload Card -- -->
  <div class="card">
    <div class="card-title">
      <div class="icon">&#x1F4C2;</div>
      Upload Workbooks
    </div>

    <div class="upload-grid">

      <!-- Rules workbook -->
      <div>
        <p style="font-size:0.8rem;font-weight:600;color:var(--text-secondary);margin-bottom:10px;text-transform:uppercase;letter-spacing:0.5px;">
          1 &middot; Validation Rules Workbook
        </p>
        <div class="drop-zone" id="rules-zone">
          <input type="file" accept=".xlsx" id="rules-input" aria-label="Upload Validation Rules Workbook" />
          <span class="dz-icon">&#x1F4CB;</span>
          <p class="dz-label">Drop file here or click to browse</p>
          <p class="dz-hint">Accepts .xlsx only</p>
          <div class="dz-filename" style="display:none;"></div>
        </div>
      </div>

      <!-- DGW data workbook -->
      <div>
        <p style="font-size:0.8rem;font-weight:600;color:var(--text-secondary);margin-bottom:10px;text-transform:uppercase;letter-spacing:0.5px;">
          2 &middot; DGW Data Workbook
        </p>
        <div class="drop-zone" id="dgw-zone">
          <input type="file" accept=".xlsx" id="dgw-input" aria-label="Upload DGW Data Workbook" />
          <span class="dz-icon">&#x1F4CA;</span>
          <p class="dz-label">Drop file here or click to browse</p>
          <p class="dz-hint">Accepts .xlsx only</p>
          <div class="dz-filename" style="display:none;"></div>
        </div>
      </div>

    </div>

    <button class="btn-run" id="run-btn" disabled>
      &#x25B6; &nbsp;Run Validation
    </button>
  </div>

  <!-- -- Progress Card -- -->
  <div class="card" id="progress-section">
    <div class="card-title">
      <div class="icon">&#x23F3;</div>
      Validation In Progress
    </div>

    <div class="progress-pct" id="progress-pct">0%</div>
    <div class="progress-track">
      <div class="progress-bar" id="progress-bar" style="width:0%;" role="progressbar" aria-valuenow="0" aria-valuemin="0" aria-valuemax="100"></div>
    </div>

    <div class="progress-current" id="progress-current">
      <div class="spinner"></div>
      <span>Loading workbooks...</span>
    </div>

    <div class="step-log" id="step-log" aria-live="polite"></div>
  </div>

  <!-- -- Results Card -- -->
  <div id="results-section">

    <!-- Summary stat cards -->
    <div class="card">
      <div class="card-title">
        <div class="icon">&#x1F4C8;</div>
        Validation Summary
      </div>

      <div class="summary-grid">
        <div class="stat-card accent">
          <div class="stat-value" id="stat-sheets">&mdash;</div>
          <div class="stat-label">Sheets Validated</div>
        </div>
        <div class="stat-card info">
          <div class="stat-value" id="stat-rows">&mdash;</div>
          <div class="stat-label">Total Data Rows</div>
        </div>
        <div class="stat-card success">
          <div class="stat-value" id="stat-success">&mdash;</div>
          <div class="stat-label">Success Rate</div>
        </div>
        <div class="stat-card danger">
          <div class="stat-value" id="stat-issues">&mdash;</div>
          <div class="stat-label">Total Issues</div>
        </div>
        <div class="stat-card danger">
          <div class="stat-value" id="stat-high">&mdash;</div>
          <div class="stat-label">&#x1F534; High</div>
        </div>
        <div class="stat-card warning">
          <div class="stat-value" id="stat-medium">&mdash;</div>
          <div class="stat-label">&#x1F7E1; Medium</div>
        </div>
        <div class="stat-card info">
          <div class="stat-value" id="stat-low">&mdash;</div>
          <div class="stat-label">&#x1F535; Low</div>
        </div>
      </div>

      <!-- Download button -->
      <div style="text-align:center;">
        <a class="btn-download" id="download-btn" href="#" style="display:none;" download>
          &#x2B07;&#xFE0F; &nbsp;Download Validation Report
        </a>
      </div>
    </div>

  </div><!-- /results-section -->

</main>

<script src="/static/app.js"></script>
</body>
</html>"""


# =======================================================================
# SECTION 6: FLASK APP
# =======================================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
REPORT_FOLDER = os.path.join(BASE_DIR, 'reports')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORT_FOLDER, exist_ok=True)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB

# In-memory job store: job_id -> job dict
jobs: dict = {}


# ---------------------------------------------------------------------------
# Static file routes
# ---------------------------------------------------------------------------

@app.route('/static/style.css')
def serve_css():
    return Response(CSS_CONTENT, mimetype='text/css')


@app.route('/static/app.js')
def serve_js():
    return Response(JS_CONTENT, mimetype='application/javascript')


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)


@app.route('/validate', methods=['POST'])
def validate():
    if 'rules_file' not in request.files or 'dgw_file' not in request.files:
        return jsonify({'error': 'Both rules_file and dgw_file are required'}), 400

    rules_file = request.files['rules_file']
    dgw_file = request.files['dgw_file']

    if not rules_file.filename.endswith('.xlsx') or not dgw_file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Both files must be .xlsx format'}), 400

    job_id = str(uuid.uuid4())

    rules_path = os.path.join(UPLOAD_FOLDER, f'{job_id}_rules.xlsx')
    dgw_path   = os.path.join(UPLOAD_FOLDER, f'{job_id}_dgw.xlsx')
    rules_file.save(rules_path)
    dgw_file.save(dgw_path)

    jobs[job_id] = {
        'status':      'running',
        'progress':    0,
        'steps':       [],
        'current':     'Starting...',
        'report_path': None,
        'summary':     None,
        'error':       None,
        'created_at':  datetime.datetime.now(),
    }

    thread = threading.Thread(
        target=_run_validation,
        args=(job_id, rules_path, dgw_path),
        daemon=True,
    )
    thread.start()

    return jsonify({'job_id': job_id}), 202


@app.route('/status/<job_id>')
def status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({'error': 'Job not found'}), 404
    return jsonify({
        'status':   job['status'],
        'progress': job['progress'],
        'current':  job['current'],
        'steps':    job['steps'],
        'summary':  job['summary'],
        'error':    job['error'],
    })


@app.route('/download/<job_id>')
def download(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({'error': 'Job not found'}), 404
    if job['status'] != 'complete':
        return jsonify({'error': 'Report not ready yet'}), 400
    report_path = job.get('report_path')
    if not report_path or not os.path.exists(report_path):
        return jsonify({'error': 'Report file not found'}), 404
    return send_file(
        report_path,
        as_attachment=True,
        download_name=os.path.basename(report_path),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ---------------------------------------------------------------------------
# Background worker
# ---------------------------------------------------------------------------

def _run_validation(job_id: str, rules_path: str, dgw_path: str):
    job = jobs[job_id]
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    report_path = os.path.join(REPORT_FOLDER, f'DGW_Validation_Report_{timestamp}.xlsx')

    def progress_cb(pct: int, msg: str):
        job['progress'] = pct
        job['current'] = msg
        if msg not in job['steps'] and pct < 100:
            job['steps'].append(msg)

    try:
        engine = ValidationEngine(rules_path, dgw_path, progress_callback=progress_cb)
        summary = engine.run(report_path)

        job['status'] = 'complete'
        job['progress'] = 100
        job['current'] = 'Complete'
        job['steps'].append('Validation complete -- report ready for download')
        job['report_path'] = report_path
        job['summary'] = summary

    except Exception as exc:
        job['status'] = 'error'
        job['error'] = str(exc)
        job['current'] = f'Error: {exc}'

    finally:
        try:
            os.remove(rules_path)
        except OSError:
            pass
        try:
            os.remove(dgw_path)
        except OSError:
            pass


# ---------------------------------------------------------------------------
# Cleanup old jobs
# ---------------------------------------------------------------------------

def _cleanup_old_jobs():
    cutoff = datetime.datetime.now() - datetime.timedelta(hours=1)
    stale = [jid for jid, j in jobs.items() if j.get('created_at', datetime.datetime.now()) < cutoff]
    for jid in stale:
        job = jobs.pop(jid, {})
        rp = job.get('report_path')
        if rp and os.path.exists(rp):
            try:
                os.remove(rp)
            except OSError:
                pass


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    threading.Timer(1.2, lambda: webbrowser.open('http://localhost:5000')).start()
    print("DGW Validation Engine starting on http://localhost:5000")
    app.run(debug=False, port=5000)
