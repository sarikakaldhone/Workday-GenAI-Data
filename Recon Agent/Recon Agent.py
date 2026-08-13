"""
Data Reconciliation Agent - Fixed 2-Sheet Report Version

- Keeps the improved UI with progress bar and live summary
- Detects DGW/Workday business header rows correctly
- Compares every matched row and every matched column
- Captures DGW-only and Workday-only keys
- Correctly classifies blank/value discrepancies
- Generates only 2 report sheets in the earlier format:
  1. <Sheet Name> - Summary
  2. <Sheet Name> - Discrepancies
- Downloads the report as: Reconciliation report.xlsx
"""

import os
import re
import time
import uuid
import tempfile
import threading
import traceback
from pathlib import Path
from difflib import SequenceMatcher
from collections import Counter, defaultdict
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, request, render_template_string, jsonify, send_file, redirect, url_for
from werkzeug.utils import secure_filename

ALLOWED_EXTENSIONS = {"xlsx", "xlsm"}
MAX_CONTENT_LENGTH = 100 * 1024 * 1024
WORK_DIR = Path(tempfile.gettempdir()) / "recon_web"
WORK_DIR.mkdir(parents=True, exist_ok=True)
HEADER_SCAN_ROWS = 15
FUZZY_MATCH_THRESHOLD = 0.82

EXPECTED_HEADER_TERMS = {
    "workerid", "employeeid", "workertype", "addresstype", "primary",
    "eventeffectivedate", "addresseffectivedate", "countryisocode",
    "addressline1", "city", "postalcode", "addressreferenceid",
    "isworkerremote", "addressline1local"
}
PRIMARY_KEY_CANDIDATES = ["Worker ID", "Employee ID", "EmployeeID", "Person ID", "Reference ID", "WID"]
SYNONYMS = {
    "employeeid": {"workerid", "personid", "referenceid", "wid"},
    "workerid": {"employeeid", "personid", "referenceid", "wid"},
    "region": {"countryregionreferenceid", "acnvrcountryregionreferenceid"},
    "countryregionreferenceid": {"region", "acnvrcountryregionreferenceid"},
    "addressusagebehavior1": {"communicationusagebehavior1"},
    "addressusagebehavior2": {"communicationusagebehavior2"},
    "addressusagebehavior3": {"communicationusagebehavior3"},
}

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_LENGTH
app.secret_key = "change-me-in-production"

JOBS = {}
JOBS_LOCK = threading.Lock()

def new_job(run_dir):
    job_id = uuid.uuid4().hex[:12]
    with JOBS_LOCK:
        JOBS[job_id] = {
            "pct": 0.0, "message": "Queued...", "started_at": time.time(), "updated_at": time.time(),
            "done": False, "error": None, "file_path": None, "download_name": None,
            "banner": None, "summary": None, "run_dir": str(run_dir)
        }
    return job_id

def set_job(job_id, pct=None, message=None):
    with JOBS_LOCK:
        if job_id not in JOBS:
            return
        j = JOBS[job_id]
        if pct is not None:
            j["pct"] = max(j["pct"], float(pct))
        if message is not None:
            j["message"] = message
        j["updated_at"] = time.time()

def set_summary(job_id, rows):
    with JOBS_LOCK:
        if job_id in JOBS:
            JOBS[job_id]["summary"] = rows
            JOBS[job_id]["updated_at"] = time.time()

def finish_job(job_id, file_path=None, download_name=None, banner=None, error=None):
    with JOBS_LOCK:
        if job_id not in JOBS:
            return
        j = JOBS[job_id]
        j["pct"] = 100.0
        j["done"] = True
        j["file_path"] = str(file_path) if file_path else None
        j["download_name"] = download_name
        j["banner"] = banner
        j["error"] = error
        j["message"] = f"Error: {error}" if error else "Complete!"
        j["updated_at"] = time.time()

PAGE_HTML = """
<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Data Reconciliation Agent</title>
<style>
body{font-family:Segoe UI,Arial,sans-serif;background:#f7f8fb;margin:0;color:#1f2937}.wrap{max-width:980px;margin:40px auto;background:#fff;border-radius:14px;padding:28px;box-shadow:0 10px 30px rgba(20,20,40,.08)}h1{margin:0 0 8px;font-size:28px}.sub{color:#6b7280;margin-bottom:24px}.grid{display:grid;grid-template-columns:1fr 1fr;gap:18px}.card{border:1px solid #e5e7eb;border-radius:12px;padding:18px;background:#fbfdff}label{font-weight:600;display:block;margin-bottom:10px}input[type=file]{width:100%}button{margin-top:22px;background:#2563eb;color:#fff;border:0;border-radius:10px;padding:12px 18px;font-size:15px;cursor:pointer}button:disabled{background:#9ca3af}.progress-box{margin-top:24px;display:none}.bar{height:20px;background:#e5e7eb;border-radius:999px;overflow:hidden}.fill{height:100%;width:0%;background:linear-gradient(90deg,#2563eb,#22c55e);transition:width .3s}.status{margin-top:10px;color:#374151}.banner{display:none;margin-top:18px;padding:14px;border-radius:10px}.ok{background:#ecfdf5;color:#065f46;border:1px solid #a7f3d0}.warn{background:#fffbeb;color:#92400e;border:1px solid #fde68a}.error{background:#fef2f2;color:#991b1b;border:1px solid #fecaca}table{width:100%;border-collapse:collapse;margin-top:18px;font-size:14px}th,td{border-bottom:1px solid #e5e7eb;padding:9px;text-align:left}th{background:#f3f4f6}.download{display:none;margin-top:18px}.download a{color:#2563eb;font-weight:700;text-decoration:none}
</style></head><body><div class="wrap"><h1>Data Reconciliation Agent</h1><div class="sub">Upload DGW source and Workday target files. The app validates headers, keys and all row-level value discrepancies.</div>
<form id="form"><div class="grid"><div class="card"><label>DGW Source File</label><input type="file" name="dgw" accept=".xlsx,.xlsm" required></div><div class="card"><label>Workday Target File</label><input type="file" name="workday" accept=".xlsx,.xlsm" required></div></div><button id="startBtn" type="submit">Start Reconciliation</button></form>
<div class="progress-box" id="progressBox"><div class="bar"><div class="fill" id="fill"></div></div><div class="status" id="status">Starting...</div></div><div class="banner" id="banner"></div><div id="summary"></div><div class="download" id="download"></div></div>
<script>
const form=document.getElementById('form'),btn=document.getElementById('startBtn'),box=document.getElementById('progressBox'),fill=document.getElementById('fill'),statusEl=document.getElementById('status'),banner=document.getElementById('banner'),summary=document.getElementById('summary'),download=document.getElementById('download');
form.addEventListener('submit',async e=>{e.preventDefault();btn.disabled=true;banner.style.display='none';summary.innerHTML='';download.style.display='none';box.style.display='block';fill.style.width='0%';statusEl.textContent='Uploading files...';const fd=new FormData(form);const res=await fetch('/start',{method:'POST',body:fd});const data=await res.json();if(!data.ok){statusEl.textContent=data.error||'Could not start.';btn.disabled=false;return;}poll(data.job_id);});
async function poll(jobId){const res=await fetch('/progress/'+jobId);const data=await res.json();fill.style.width=(data.pct||0)+'%';statusEl.textContent=(data.pct||0)+'% - '+(data.message||'Processing...');if(data.summary&&data.summary.length){let html='<table><thead><tr><th>Metric</th><th>Value</th></tr></thead><tbody>';for(const r of data.summary)html+=`<tr><td>${r[0]}</td><td>${r[1]}</td></tr>`;html+='</tbody></table>';summary.innerHTML=html;}if(data.banner){banner.className='banner '+data.banner.kind;banner.innerHTML=data.banner.html;banner.style.display='block';}if(data.done){btn.disabled=false;if(!data.error){download.innerHTML=`<a href="/download/${jobId}">Download Reconciliation report.xlsx</a>`;download.style.display='block';}return;}setTimeout(()=>poll(jobId),800);}
</script></body></html>
"""

def allowed(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def norm_header(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())

def display_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)

def norm_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).replace("\u00a0", " ").strip()

def fuzzy_sim(a, b):
    na, nb = norm_header(a), norm_header(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    if nb in SYNONYMS.get(na, set()) or na in SYNONYMS.get(nb, set()):
        return 0.95
    if na in nb or nb in na:
        return max(0.86, SequenceMatcher(None, na, nb).ratio())
    return SequenceMatcher(None, na, nb).ratio()

def is_likely_id(value):
    return bool(re.fullmatch(r"\d{4,}", norm_value(value)))

def row_values(ws, row_num):
    return [ws.cell(row_num, c).value for c in range(1, ws.max_column + 1)]

def detect_header_row(ws):
    best = None
    for r in range(1, min(HEADER_SCAN_ROWS, ws.max_row) + 1):
        vals = row_values(ws, r)
        non_empty = [v for v in vals if norm_value(v)]
        if len(non_empty) < 3:
            continue
        normalized = [norm_header(v) for v in non_empty]
        expected_hits = sum(1 for h in normalized if h in EXPECTED_HEADER_TERMS)
        exact_worker_id = 1 if any(h in {"workerid", "employeeid"} for h in normalized) else 0
        exact_worker_type = 1 if "workertype" in normalized else 0
        next_looks_data = 1 if is_likely_id(ws.cell(r + 1, 1).value if r + 1 <= ws.max_row else None) else 0
        required_penalty = sum(1 for h in normalized if h in {"required", "optional", "text", "truefalse"})
        slash_penalty = sum(1 for v in vals if isinstance(v, str) and "/" in v)
        score = expected_hits * 10 + exact_worker_id * 25 + exact_worker_type * 10 + next_looks_data * 40 + len(non_empty) * 0.2 - required_penalty * 10 - slash_penalty * 1.5
        if best is None or score > best[0]:
            best = (score, r)
    if not best:
        raise ValueError(f"Could not detect header row for sheet '{ws.title}'.")
    return best[1]

def get_headers(ws, header_row):
    headers, seen = [], defaultdict(int)
    for c in range(1, ws.max_column + 1):
        name = str(ws.cell(header_row, c).value).strip() if ws.cell(header_row, c).value is not None else f"Blank Column {c}"
        seen[name] += 1
        if seen[name] > 1:
            name = f"{name} [{seen[name]}]"
        headers.append(name)
    return headers

def read_sheet_records(path):
    wb = load_workbook(path, data_only=True)
    best_sheet, best_score, best_header = None, -1, None
    for ws in wb.worksheets:
        if ws.max_row < 2 or ws.max_column < 2:
            continue
        try:
            hrow = detect_header_row(ws)
            headers = get_headers(ws, hrow)
            score = sum(1 for h in headers if norm_header(h) in EXPECTED_HEADER_TERMS)
            if any(norm_header(h) in {"workerid", "employeeid"} for h in headers):
                score += 20
            if score > best_score:
                best_sheet, best_score, best_header = ws.title, score, hrow
        except Exception:
            pass
    if not best_sheet:
        wb.close()
        raise ValueError(f"No business sheet with a detectable header was found in {os.path.basename(path)}")
    ws = wb[best_sheet]
    headers = get_headers(ws, best_header)
    records = []
    for r in range(best_header + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(norm_value(v) for v in vals):
            records.append({"__row__": r, **dict(zip(headers, vals))})
    wb.close()
    return {"path": path, "sheet": best_sheet, "header_row": best_header, "headers": headers, "records": records}

def match_columns(dgw_headers, wd_headers):
    mappings, used_wd = [], set()
    for dh in dgw_headers:
        if dh.startswith("Blank Column"):
            continue
        best_conf, best_wd = 0.0, None
        for wh in wd_headers:
            if wh in used_wd or wh.startswith("Blank Column"):
                continue
            sim = fuzzy_sim(dh, wh)
            if sim > best_conf:
                best_conf, best_wd = sim, wh
        if best_wd and best_conf >= FUZZY_MATCH_THRESHOLD:
            mappings.append({"dgw_column": dh, "workday_column": best_wd, "confidence": round(best_conf, 3), "compared": "Yes", "reason": "Matched"})
            used_wd.add(best_wd)
        else:
            mappings.append({"dgw_column": dh, "workday_column": "", "confidence": round(best_conf, 3), "compared": "No", "reason": "No Workday column met threshold"})
    wd_unmatched = [h for h in wd_headers if h not in used_wd and not h.startswith("Blank Column")]
    return mappings, wd_unmatched

def find_primary_key(headers):
    for candidate in PRIMARY_KEY_CANDIDATES:
        for h in headers:
            if norm_header(h) == norm_header(candidate):
                return h
    for h in headers:
        nh = norm_header(h)
        if "workerid" in nh or "employeeid" in nh or nh in {"wid", "referenceid"}:
            return h
    raise ValueError("Could not identify a primary key column such as Worker ID or Employee ID.")

def build_key_index(records, key_col):
    idx = defaultdict(list)
    for rec in records:
        key = norm_value(rec.get(key_col))
        if key:
            idx[key].append(rec)
    return idx

# ----------------------------------------------------------------------
# Discrepancy Classification (per "Discrepancy Classification Rules 2.0")
# Priority order for VALUE-level comparison of a matched key:
#   1. Missing Value  (one side blank, the other populated)
#   2. Case Difference (values equal ignoring capitalization only)
#   3. Formatting Difference (logically equal after normalization:
#         leading zeros, thousands separators, date formats, whitespace)
#   4. Value Mismatch (truly different even after normalization)
# Record-level types (Duplicate Key, Missing Record, Extra Record,
# Exact Match) are handled in reconcile().
# ----------------------------------------------------------------------

# Date formats we attempt to canonicalize before comparison.
_DATE_FORMATS = (
    "%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y", "%m-%d-%Y", "%m/%d/%Y",
    "%Y.%m.%d", "%d.%m.%Y", "%d-%b-%Y", "%d %b %Y", "%b %d, %Y",
    "%d-%B-%Y", "%d %B %Y", "%B %d, %Y",
    "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%m/%d/%Y %H:%M:%S",
    "%d-%m-%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S",
)

def _num_norm(s):
    """Return a canonical numeric string if s is a formatted number, else None.
    Handles thousands separators and leading zeros (e.g. '00123' -> '123',
    '1,000' -> '1000', '1,000.50' -> '1000.5')."""
    t = str(s).strip().replace(",", "").replace("\u00a0", "")
    if re.fullmatch(r"[+-]?\d+", t):
        return str(int(t))
    if re.fullmatch(r"[+-]?\d*\.\d+", t):
        return format(float(t), ".10g")
    return None

def _date_norm(s):
    """Return an ISO date (YYYY-MM-DD) if s parses as a date, else None."""
    t = str(s).strip()
    if not re.search(r"\d", t):
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(t, fmt).date().isoformat()
        except ValueError:
            continue
    return None

def _loose_norm(s):
    """Collapse internal whitespace so values differing only by spacing match."""
    return re.sub(r"\s+", " ", str(s).strip())

def classify_difference(dgw_value, wd_value):
    d, w = norm_value(dgw_value), norm_value(wd_value)

    # 1. Missing Value (blank on exactly one side)
    if d and not w:
        return "Missing Value (Workday Blank)"
    if w and not d:
        return "Missing Value (DGW Blank)"
    if not d and not w:
        return None  # both blank -> not a discrepancy

    if d == w:
        return None  # identical -> not a discrepancy

    # 2. Case Difference (equal ignoring capitalization only)
    if d.lower() == w.lower():
        return "Case Difference"

    # 3. Formatting Difference - numeric (leading zeros / thousands separators)
    dn, wn = _num_norm(d), _num_norm(w)
    if dn is not None and wn is not None and dn == wn:
        return "Formatting Difference"

    # 3. Date Format Difference - different date representations
    dd, wdt = _date_norm(d), _date_norm(w)
    if dd is not None and wdt is not None and dd == wdt:
        return "Date Format Difference"

    # 3. Formatting Difference - whitespace / spacing only
    if _loose_norm(d).lower() == _loose_norm(w).lower() and _loose_norm(d):
        return "Formatting Difference"

    # 4. True data mismatch
    return "Value Mismatch"

# Severity assigned per discrepancy type.
SEVERITY_BY_TYPE = {
    "Duplicate Key": "High",
    "Missing Record": "High",
    "Extra Record": "High",
    "Value Mismatch": "High",
    "Missing Value (Workday Blank)": "Medium",
    "Missing Value (DGW Blank)": "Medium",
    "Date Format Difference": "Low",
    "Formatting Difference": "Low",
    "Case Difference": "Low",
}

def severity_for(dtype):
    return SEVERITY_BY_TYPE.get(dtype, "Medium")

def reconcile(dgw_path, wd_path, out_path, on_progress=None):
    if on_progress: on_progress(10, "Reading DGW workbook...")
    dgw = read_sheet_records(dgw_path)
    if on_progress: on_progress(25, "Reading Workday workbook...")
    wd = read_sheet_records(wd_path)
    dgw_key, wd_key = find_primary_key(dgw["headers"]), find_primary_key(wd["headers"])
    if on_progress: on_progress(40, "Matching columns...")
    mappings, wd_unmatched = match_columns(dgw["headers"], wd["headers"])
    compared_mappings = [m for m in mappings if m["compared"] == "Yes"]
    dgw_index, wd_index = build_key_index(dgw["records"], dgw_key), build_key_index(wd["records"], wd_key)
    dgw_dupes = {k: v for k, v in dgw_index.items() if len(v) > 1}
    wd_dupes = {k: v for k, v in wd_index.items() if len(v) > 1}
    discrepancies, key_audit = [], []
    all_keys = sorted(set(dgw_index) | set(wd_index), key=lambda x: (not x.isdigit(), int(x) if x.isdigit() else x))
    total_keys = max(len(all_keys), 1)
    if on_progress: on_progress(55, "Comparing keys and values...")
    for i, key in enumerate(all_keys, start=1):
        if i % 10 == 0 and on_progress:
            on_progress(55 + (i / total_keys) * 30, f"Comparing key {i} of {total_keys}...")
        dgw_rows, wd_rows = dgw_index.get(key, []), wd_index.get(key, [])

        # Rule 1 (highest priority): Duplicate Key - key repeats within a dataset.
        # Classified as Duplicate Key even if other column values differ.
        if len(dgw_rows) > 1 or len(wd_rows) > 1:
            sources = []
            if len(dgw_rows) > 1: sources.append(f"DGW x{len(dgw_rows)}")
            if len(wd_rows) > 1: sources.append(f"Workday x{len(wd_rows)}")
            discrepancies.append({"Source Row (DGW)": dgw_rows[0].get("__row__", "") if dgw_rows else "", "Primary Key": key, "Column Name": "Record", "DGW Original Value": f"{len(dgw_rows)} row(s)", "Workday Original Value": f"{len(wd_rows)} row(s)", "Discrepancy Type": "Duplicate Key", "Difference Details": f"Primary key '{key}' appears more than once ({', '.join(sources)}).", "Severity": severity_for("Duplicate Key"), "Rule / Matching Notes": "Rule 1: Duplicate primary key within a dataset", "Workday Row Ref": wd_rows[0].get("__row__", "") if wd_rows else ""})
            key_audit.append([key, "Duplicate Key", dgw_rows[0].get("__row__", "") if dgw_rows else "", wd_rows[0].get("__row__", "") if wd_rows else ""])
            continue

        # Rule 3: Extra Record - key exists in target (Workday) but not source (DGW).
        if not dgw_rows:
            for wd_rec in wd_rows:
                discrepancies.append({"Source Row (DGW)": "", "Primary Key": key, "Column Name": "Record", "DGW Original Value": "", "Workday Original Value": "Present", "Discrepancy Type": "Extra Record", "Difference Details": f"Workday has key '{key}' but DGW does not.", "Severity": severity_for("Extra Record"), "Rule / Matching Notes": "Rule 3: Record exists in Workday (target) but not DGW (source)", "Workday Row Ref": wd_rec.get("__row__", "")})
                key_audit.append([key, "Extra Record", "", wd_rec.get("__row__", "")])
            continue

        # Rule 2: Missing Record - key exists in source (DGW) but not target (Workday).
        if not wd_rows:
            for dgw_rec in dgw_rows:
                discrepancies.append({"Source Row (DGW)": dgw_rec.get("__row__", ""), "Primary Key": key, "Column Name": "Record", "DGW Original Value": "Present", "Workday Original Value": "", "Discrepancy Type": "Missing Record", "Difference Details": f"DGW has key '{key}' but Workday does not.", "Severity": severity_for("Missing Record"), "Rule / Matching Notes": "Rule 2: Record exists in DGW (source) but not Workday (target)", "Workday Row Ref": ""})
                key_audit.append([key, "Missing Record", dgw_rec.get("__row__", ""), ""])
            continue
        dgw_rec, wd_rec = dgw_rows[0], wd_rows[0]
        key_audit.append([key, "Matched", dgw_rec.get("__row__", ""), wd_rec.get("__row__", "")])
        for m in compared_mappings:
            dh, wh = m["dgw_column"], m["workday_column"]
            if norm_header(dh) == norm_header(dgw_key) and norm_header(wh) == norm_header(wd_key):
                continue
            d_orig, w_orig = dgw_rec.get(dh), wd_rec.get(wh)
            if norm_value(d_orig) != norm_value(w_orig):
                dtype = classify_difference(d_orig, w_orig)
                if not dtype:
                    continue  # normalized-equal that resolves to no discrepancy
                discrepancies.append({"Source Row (DGW)": dgw_rec.get("__row__", ""), "Primary Key": key, "Column Name": dh, "DGW Original Value": display_value(d_orig), "Workday Original Value": display_value(w_orig), "Discrepancy Type": dtype, "Difference Details": f'DGW: "{display_value(d_orig)}" vs Workday: "{display_value(w_orig)}"', "Severity": severity_for(dtype), "Rule / Matching Notes": f"Compared {dh} -> {wh} at confidence {m['confidence']}", "Workday Row Ref": wd_rec.get("__row__", "")})
    if on_progress: on_progress(90, "Writing report...")
    write_report(out_path, dgw, wd, mappings, wd_unmatched, discrepancies, key_audit, dgw_dupes, wd_dupes, dgw_key, wd_key)
    return {"dgw_sheet": dgw["sheet"], "wd_sheet": wd["sheet"], "dgw_header_row": dgw["header_row"], "wd_header_row": wd["header_row"], "dgw_rows": len(dgw["records"]), "wd_rows": len(wd["records"]), "matched_rows": len(set(dgw_index) & set(wd_index)), "discrepancies": len(discrepancies), "dgw_only_keys": len(set(dgw_index) - set(wd_index)), "wd_only_keys": len(set(wd_index) - set(dgw_index)), "columns_compared": max(0, len(compared_mappings) - 1), "type_counts": Counter(d["Discrepancy Type"] for d in discrepancies)}

def write_report(out_path, dgw, wd, mappings, wd_unmatched, discrepancies, key_audit, dgw_duplicate_keys, wd_duplicate_keys, dgw_key, wd_key):
    wb = Workbook()
    wb.remove(wb.active)
    base = str(dgw["sheet"])[:22] if dgw.get("sheet") else "Reconciliation"
    ws = wb.create_sheet(f"{base} - Summary")
    ws2 = wb.create_sheet(f"{base} - Discrepancies")
    title_fill, section_fill, header_fill, light_fill = PatternFill("solid", fgColor="1F4E78"), PatternFill("solid", fgColor="D9EAF7"), PatternFill("solid", fgColor="1F4E78"), PatternFill("solid", fgColor="EAF2F8")
    white_font, bold_font = Font(bold=True, color="FFFFFF"), Font(bold=True, color="000000")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_matched = sum(1 for r in key_audit if r[1] == "Matched")
    type_counts = Counter(d["Discrepancy Type"] for d in discrepancies)
    compared = [m for m in mappings if m["compared"] == "Yes" and norm_header(m["dgw_column"]) != norm_header(dgw_key)]
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:J1"); ws["A1"] = "DGW vs Workday Reconciliation Report"; ws["A1"].font = Font(bold=True, size=16, color="FFFFFF"); ws["A1"].fill = title_fill; ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:J2"); ws["A2"] = f"Generated: {time.strftime('%Y-%m-%d %H:%M')}   |   DGW: {Path(dgw['path']).name}   |   Workday: {Path(wd['path']).name}"; ws["A2"].fill = light_fill; ws["A2"].alignment = Alignment(horizontal="center")
    r = 4; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10); ws.cell(r,1).value = "Reconciliation Overview"; ws.cell(r,1).font = bold_font; ws.cell(r,1).fill = section_fill
    overview = [["DGW Sheet Name", dgw["sheet"]], ["Matched Workday Sheet", wd["sheet"]], ["Column Match Confidence", "100%" if not wd_unmatched and not any(m["compared"] == "No" for m in mappings) else "Partial"], ["Header Row (DGW)", dgw["header_row"]], ["Total DGW Data Rows", len(dgw["records"])], ["Total Matched Rows", total_matched], ["Total Discrepancies", len(discrepancies)], ["Discrepancy Rate %", f"{(len(discrepancies) / total_matched) if total_matched else 0:.1%}"], ["Unique Keys in DGW only (not in WD)", sum(1 for row in key_audit if row[1] == "Missing in Workday")], ["Unique Keys in Workday only (not in DGW)", sum(1 for row in key_audit if row[1] == "Missing in DGW")], ["Columns Matched (DGW → WD)", len(compared)], ["Run Date", time.strftime('%Y-%m-%d')]]
    r += 1
    for label, value in overview:
        ws.cell(r,1).value = label; ws.cell(r,2).value = value; ws.cell(r,1).font = bold_font; ws.cell(r,1).fill = light_fill; ws.cell(r,1).border = border; ws.cell(r,2).border = border; r += 1
    r += 1
    for c, h in enumerate(["Discrepancy Type", "Count", "% of Matched Rows"], start=1):
        ws.cell(r,c).value = h; ws.cell(r,c).font = white_font; ws.cell(r,c).fill = header_fill; ws.cell(r,c).border = border
    r += 1
    for dtype, cnt in type_counts.items():
        ws.cell(r,1).value = dtype; ws.cell(r,2).value = cnt; ws.cell(r,3).value = f"{(cnt / total_matched) if total_matched else 0:.1%}"
        for c in range(1,4): ws.cell(r,c).border = border
        r += 1
    r += 2
    for c, h in enumerate(["Column Name", "DGW → WD Mapping", "Discrepancy Count", "% of Matched Rows"], start=1):
        ws.cell(r,c).value = h; ws.cell(r,c).font = white_font; ws.cell(r,c).fill = header_fill; ws.cell(r,c).border = border
    r += 1
    col_counts = Counter(d["Column Name"] for d in discrepancies if d["Column Name"] != "Record")
    lookup = {m["dgw_column"]: m for m in mappings}
    for col, cnt in col_counts.items():
        m = lookup.get(col, {}); wd_col = m.get("workday_column", col); conf = m.get("confidence", "")
        mapping_text = f"{col} → {wd_col} ({int(float(conf)*100)}%)" if conf != "" else f"{col} → {wd_col}"
        ws.cell(r,1).value = col; ws.cell(r,2).value = mapping_text; ws.cell(r,3).value = cnt; ws.cell(r,4).value = f"{(cnt / total_matched) if total_matched else 0:.1%}"
        for c in range(1,5): ws.cell(r,c).border = border
        r += 1
    for c in range(1,11): ws.column_dimensions[get_column_letter(c)].width = 28 if c <= 4 else 14
    ws2.sheet_view.showGridLines = False
    headers = ["Source Row\n(DGW)", "Primary Key", "Column Name", "DGW Original Value", "Workday Original Value", "Discrepancy Type", "Difference Details", "Severity", "Rule / Matching Notes", "Workday\nRow Ref"]
    ws2.append(headers)
    for d in discrepancies:
        ws2.append([d.get("Source Row (DGW)", ""), d.get("Primary Key", ""), d.get("Column Name", ""), d.get("DGW Original Value", ""), d.get("Workday Original Value", ""), d.get("Discrepancy Type", ""), d.get("Difference Details", ""), d.get("Severity", ""), d.get("Rule / Matching Notes", ""), d.get("Workday Row Ref", "")])
    for cell in ws2[1]: cell.font = white_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
    # Severity text colors: High=red, Medium=orange, Low=green
    severity_colors = {"High": "C00000", "Medium": "ED7D31", "Low": "00B050"}
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        sev = row[7].value
        if sev in severity_colors:
            row[7].font = Font(bold=True, color=severity_colors[sev])
    for i, width in enumerate([14,16,26,28,28,26,55,12,38,14], start=1): ws2.column_dimensions[get_column_letter(i)].width = width
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws2.max_row}"
    wb.save(out_path)

def worker(job_id, dgw_path, wd_path, run_dir):
    try:
        set_job(job_id, 5, "Starting reconciliation...")
        out_path = Path(run_dir) / "Reconciliation report.xlsx"
        stats = reconcile(dgw_path, wd_path, out_path, on_progress=lambda p, m: set_job(job_id, p, m))
        set_summary(job_id, [["DGW Sheet", stats["dgw_sheet"]], ["Workday Sheet", stats["wd_sheet"]], ["DGW Header Row", stats["dgw_header_row"]], ["Workday Header Row", stats["wd_header_row"]], ["DGW Rows", stats["dgw_rows"]], ["Workday Rows", stats["wd_rows"]], ["Matched Keys", stats["matched_rows"]], ["Total Discrepancies", stats["discrepancies"]], ["DGW-only Keys", stats["dgw_only_keys"]], ["Workday-only Keys", stats["wd_only_keys"]], ["Columns Compared", stats["columns_compared"]]])
        banner = {"kind": "warn", "html": f"Reconciliation complete. Found <b>{stats['discrepancies']}</b> discrepancies. Download the report for full detail."} if stats["discrepancies"] else {"kind": "ok", "html": "Reconciliation complete. No discrepancies found."}
        finish_job(job_id, file_path=out_path, download_name="Reconciliation report.xlsx", banner=banner)
    except Exception as exc:
        traceback.print_exc()
        finish_job(job_id, error=str(exc), banner={"kind": "error", "html": str(exc)})

@app.route("/", methods=["GET"])
def index():
    return render_template_string(PAGE_HTML)

@app.route("/start", methods=["POST"])
def start():
    dgw_file, wd_file = request.files.get("dgw"), request.files.get("workday")
    if not dgw_file or not wd_file or not dgw_file.filename or not wd_file.filename:
        return jsonify(ok=False, error="Please choose BOTH files.")
    if not (allowed(dgw_file.filename) and allowed(wd_file.filename)):
        return jsonify(ok=False, error="Only .xlsx or .xlsm files are supported.")
    run_dir = WORK_DIR / uuid.uuid4().hex; run_dir.mkdir(parents=True, exist_ok=True)
    dgw_path, wd_path = run_dir / secure_filename(dgw_file.filename), run_dir / secure_filename(wd_file.filename)
    dgw_file.save(dgw_path); wd_file.save(wd_path)
    job_id = new_job(run_dir)
    threading.Thread(target=worker, args=(job_id, dgw_path, wd_path, run_dir), daemon=True).start()
    return jsonify(ok=True, job_id=job_id)

@app.route("/progress/<job_id>", methods=["GET"])
def progress(job_id):
    with JOBS_LOCK:
        j = JOBS.get(job_id)
        if not j: return jsonify(done=True, error="Unknown job")
        return jsonify(pct=round(j["pct"], 1), message=j["message"], done=j["done"], error=j["error"], banner=j.get("banner"), summary=j.get("summary"))

@app.route("/download/<job_id>", methods=["GET"])
def download(job_id):
    with JOBS_LOCK:
        j = JOBS.get(job_id)
        if not j or not j.get("file_path"):
            return redirect(url_for("index"))
        fp, name = j["file_path"], j.get("download_name") or "Reconciliation report.xlsx"
    return send_file(fp, as_attachment=True, download_name=name)

if __name__ == "__main__":
    print("=" * 68)
    print(" Data Reconciliation Agent - Fixed 2-sheet report")
    print(" Report output: Reconciliation report.xlsx")
    print("=" * 68)
    app.run(host="127.0.0.1", port=5000, debug=False, threaded=True)
